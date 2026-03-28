"""
api/routes/excel_dashboard.py — Excel do Dashboard (DS + cidades + regiões)
"""
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from api.deps import get_supabase, get_current_user
from api.limiter import limiter
from api.routes.excel_base import (
    _HF, _AF, _RF, _GF, _HFNT, _BFNT, _CTR, _LFT, _BRD,
    _titulo_aba, _write_grouped, _to_stream,
)
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

router = APIRouter()


@router.get("/dashboard/{data_ref}")
@limiter.limit("20/minute")
def excel_dashboard(request: Request, data_ref: str, user: dict = Depends(get_current_user)):
    sb = get_supabase()

    q = sb.table("expedicao_diaria").select("*").eq("data_ref", data_ref)
    if user["bases"]: q = q.in_("scan_station", user["bases"])
    dia = pd.DataFrame(q.execute().data or [])
    if dia.empty: raise HTTPException(404, "Sem dados")

    qc = sb.table("expedicao_cidades").select("*").eq("data_ref", data_ref)
    if user["bases"]: qc = qc.in_("scan_station", user["bases"])
    cid = pd.DataFrame(qc.execute().data or [])

    dia = dia.sort_values("recebido", ascending=False)
    wb = Workbook()

    # ── Aba Consolidado Geral (DS + cidades) ──────────────────
    ws = wb.active; ws.title = "Consolidado_Geral"
    _titulo_aba(ws, f"Consolidado Geral — {data_ref}", 7)
    _write_grouped(ws, dia, cid, start_row=3)

    # ── Abas por região (DS + cidades) ────────────────────────
    for regiao, label in [("capital", "Capital"), ("metropolitan", "Metropolitan"), ("countryside", "Countryside")]:
        df_r = dia[dia["region"].str.lower() == regiao]
        if df_r.empty: continue
        cid_r = cid[cid["scan_station"].isin(df_r["scan_station"])] if not cid.empty else pd.DataFrame()
        ws_r = wb.create_sheet(label)
        _titulo_aba(ws_r, label, 7)
        _write_grouped(ws_r, df_r, cid_r, start_row=3)

    # ── Aba Resumo Dinâmico (regiões lado a lado) ─────────────
    ws_res = wb.create_sheet("Resumo_Dinamico")
    ws_res.sheet_view.showGridLines = False

    REGIOES = [("capital", "Capital"), ("metropolitan", "Metropolitan"), ("countryside", "Countryside")]
    COLS_RES = ["DS", "recebido no DS", "em rota de entrega", "Total Geral", "Taxa de Expedicao"]

    total_cols = len(COLS_RES) * 3 + 2
    ws_res.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws_res.cell(1, 1, f"Resumo Consolidado — {data_ref}")
    t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill = _HF; t.alignment = _CTR; ws_res.row_dimensions[1].height = 30

    col_start = 1
    for regiao, label in REGIOES:
        df_r = dia[dia["region"].str.lower() == regiao].copy()
        if df_r.empty: continue

        rec  = int(df_r["recebido"].sum())
        exp  = int(df_r["expedido"].sum())
        taxa = round(exp / rec, 4) if rec else 0

        ws_res.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_start + len(COLS_RES) - 1)
        h = ws_res.cell(2, col_start, f"{label}   |   Rec: {rec:,}   Exp: {exp:,}   Taxa: {taxa:.1%}")
        h.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        h.fill = _HF; h.alignment = _CTR; ws_res.row_dimensions[2].height = 24

        for ci, cn in enumerate(COLS_RES, col_start):
            c = ws_res.cell(3, ci, cn)
            c.font = _HFNT; c.fill = _HF; c.alignment = _CTR; c.border = _BRD
        ws_res.row_dimensions[3].height = 22

        df_r = df_r.sort_values("recebido", ascending=False)
        for ri, (_, r) in enumerate(df_r.iterrows(), 4):
            taxa_ds = float(r.get("taxa_exp", 0) or 0)
            meta_ds = float(r.get("meta", 0.5) or 0.5)
            alt = ri % 2 == 0
            vals = [r["scan_station"], int(r.get("recebido", 0)), int(r.get("expedido", 0)),
                    int(r.get("recebido", 0)), taxa_ds]
            for ci, val in enumerate(vals, col_start):
                c = ws_res.cell(ri, ci, val)
                c.font = _BFNT; c.border = _BRD; c.alignment = _CTR
                if alt: c.fill = _AF
                if ci == col_start: c.alignment = Alignment(horizontal="left", vertical="center")
                if ci == col_start + 4:
                    c.number_format = "0.0%"
                    c.fill = _RF if taxa_ds < meta_ds else _GF
                    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                elif ci in (col_start + 1, col_start + 2, col_start + 3):
                    c.number_format = "#,##0"

        last_r = 4 + len(df_r)
        totals = ["Total Geral", rec, exp, rec, taxa]
        for ci, val in enumerate(totals, col_start):
            c = ws_res.cell(last_r, ci, val)
            c.fill = PatternFill("solid", fgColor="BDD7EE")
            c.font = Font(name="Arial", bold=True, size=10, color="1F3864")
            c.border = _BRD; c.alignment = _CTR
            if ci == col_start: c.alignment = Alignment(horizontal="left", vertical="center")
            if ci == col_start + 4: c.number_format = "0.0%"
            elif ci in (col_start + 1, col_start + 2, col_start + 3): c.number_format = "#,##0"

        for ci2, w in enumerate([22, 16, 18, 14, 16], col_start):
            ws_res.column_dimensions[get_column_letter(ci2)].width = w

        col_start += len(COLS_RES) + 1

    ws_res.freeze_panes = "A4"

    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Dashboard_{data_ref}.xlsx"},
    )

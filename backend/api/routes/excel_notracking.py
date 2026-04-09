"""
api/routes/excel_notracking.py — Excel No Tracking (断更)
"""
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from api.deps import get_db, get_current_user
from api.limiter import limiter
from api.routes.excel_base import (
    _HF, _HFNT, _BFNT, _CTR, _LFT, _BRD, _AF,
    _titulo_aba, _write_header, _write_data, _auto_width, _to_stream,
)
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

router = APIRouter()


@router.get("/notracking/{upload_id}")
@limiter.limit("10/minute")
def excel_notracking(
    upload_id: int, request: Request,
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    up_row = db.execute(
        text("SELECT * FROM notracking_uploads WHERE id = :id"), {"id": upload_id}
    ).mappings().first()
    if not up_row:
        raise HTTPException(404, "Upload não encontrado.")
    up = dict(up_row)
    data_ref = up.get("data_ref", "")

    por_ds_rows = db.execute(
        text("SELECT station, supervisor, regional, total, valor_total, total_7d_mais FROM notracking_por_ds WHERE upload_id = :uid ORDER BY total DESC"),
        {"uid": upload_id}
    ).mappings().all()
    por_ds = [dict(r) for r in por_ds_rows]

    por_sup_rows = db.execute(
        text("SELECT supervisor, total, total_7d_mais FROM notracking_por_sup WHERE upload_id = :uid ORDER BY total DESC"),
        {"uid": upload_id}
    ).mappings().all()
    por_sup = [dict(r) for r in por_sup_rows]

    por_sta_rows = db.execute(
        text("SELECT status, total FROM notracking_por_status WHERE upload_id = :uid ORDER BY total DESC"),
        {"uid": upload_id}
    ).mappings().all()
    por_sta = [dict(r) for r in por_sta_rows]

    por_faixa_rows = db.execute(
        text("SELECT faixa, total FROM notracking_por_faixa WHERE upload_id = :uid ORDER BY faixa"),
        {"uid": upload_id}
    ).mappings().all()
    por_faixa = [dict(r) for r in por_faixa_rows]

    wb = Workbook()

    # ── Aba "Por DS" ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Por DS"
    _titulo_aba(ws, f"No Tracking — Por DS — {data_ref}", 8)
    _write_header(ws, ["#", "DS", "Supervisor", "Regional", "Total", "≥7 Dias", "% ≥7d", "Valor em Risco"], 2)
    rows = []
    grand = sum(d["total"] for d in por_ds) or 1
    for i, d in enumerate(por_ds, 1):
        pct = round(d["total_7d_mais"] / d["total"] * 100, 1) if d["total"] else 0
        rows.append({
            "#": i,
            "DS": d["station"],
            "Supervisor": d["supervisor"] or "—",
            "Regional": d["regional"] or "—",
            "Total": d["total"],
            "≥7 Dias": d["total_7d_mais"],
            "% ≥7d": f"{pct}%",
            "Valor em Risco": d["valor_total"],
        })
    df_ds = pd.DataFrame(rows)
    _write_data(ws, df_ds, 3)
    _auto_width(ws)
    ws.freeze_panes = "A3"

    # ── Aba "Por Supervisor" ──────────────────────────────────
    ws2 = wb.create_sheet("Por Supervisor")
    _titulo_aba(ws2, f"No Tracking — Por Supervisor — {data_ref}", 4)
    _write_header(ws2, ["Supervisor", "Total", "≥7 Dias", "% ≥7d"], 2)
    rows2 = []
    for d in por_sup:
        pct = round(d["total_7d_mais"] / d["total"] * 100, 1) if d["total"] else 0
        rows2.append({
            "Supervisor": d["supervisor"],
            "Total": d["total"],
            "≥7 Dias": d["total_7d_mais"],
            "% ≥7d": f"{pct}%",
        })
    df_sup = pd.DataFrame(rows2) if rows2 else pd.DataFrame(columns=["Supervisor", "Total", "≥7 Dias", "% ≥7d"])
    _write_data(ws2, df_sup, 3)
    _auto_width(ws2)
    ws2.freeze_panes = "A3"

    # ── Aba "Por Status" ──────────────────────────────────────
    ws3 = wb.create_sheet("Por Status")
    _titulo_aba(ws3, f"No Tracking — Por Status — {data_ref}", 2)
    _write_header(ws3, ["Último Status", "Total"], 2)
    df_sta = pd.DataFrame(por_sta if por_sta else [], columns=["status", "total"])
    if por_sta:
        df_sta.columns = ["Último Status", "Total"]
    else:
        df_sta = pd.DataFrame(columns=["Último Status", "Total"])
    _write_data(ws3, df_sta, 3)
    _auto_width(ws3)
    ws3.freeze_panes = "A3"

    # ── Aba "Por Faixa de Aging" ──────────────────────────────
    ws4 = wb.create_sheet("Por Faixa de Aging")
    _titulo_aba(ws4, f"No Tracking — Por Faixa de Aging — {data_ref}", 2)
    _write_header(ws4, ["Faixa", "Total"], 2)
    df_faixa = pd.DataFrame(por_faixa if por_faixa else [], columns=["faixa", "total"])
    if por_faixa:
        df_faixa.columns = ["Faixa", "Total"]
    else:
        df_faixa = pd.DataFrame(columns=["Faixa", "Total"])
    _write_data(ws4, df_faixa, 3)
    _auto_width(ws4)
    ws4.freeze_panes = "A3"

    fname = f"NoTracking_{data_ref}.xlsx"
    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )

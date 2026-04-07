"""
api/routes/excel_reclamacoes.py — Excel de Reclamações (TOP Ofensores)
"""
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from api.deps import get_db, get_current_user
from api.limiter import limiter
from api.routes.excel_base import _CTR, _BRD, _to_stream
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

router = APIRouter()


@router.get("/reclamacoes/{upload_id}")
@limiter.limit("20/minute")
def excel_reclamacoes(request: Request, upload_id: int, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    up_row = db.execute(
        text("SELECT * FROM reclamacoes_uploads WHERE id = :id"), {"id": upload_id}
    ).mappings().first()
    if not up_row:
        raise HTTPException(404, "Upload não encontrado")
    u = dict(up_row)

    r_sup = pd.DataFrame([dict(r) for r in db.execute(
        text("SELECT * FROM reclamacoes_por_supervisor WHERE upload_id = :uid"), {"uid": upload_id}
    ).mappings().all()])
    r_sta = pd.DataFrame([dict(r) for r in db.execute(
        text("SELECT * FROM reclamacoes_por_station WHERE upload_id = :uid"), {"uid": upload_id}
    ).mappings().all()])
    top_all = pd.DataFrame([dict(r) for r in db.execute(
        text("SELECT * FROM reclamacoes_top5 WHERE upload_id = :uid ORDER BY total DESC"), {"uid": upload_id}
    ).mappings().all()])

    inativos_rows = db.execute(
        text("SELECT id_motorista FROM motoristas_status WHERE ativo = false")
    ).mappings().all()
    inativos = {r["id_motorista"] for r in inativos_rows}
    top5 = top_all[~top_all["motorista"].isin(inativos)].head(5) if not top_all.empty else top_all

    semana_ref   = u.get("semana_ref", 0)
    week_cols_sta = sorted([c for c in r_sta.columns if c.startswith("week_")], reverse=True) if not r_sta.empty else []
    week_cols_sup = sorted([c for c in r_sup.columns if c.startswith("week_")], reverse=True) if not r_sup.empty else []

    def _week_label(col): return f"Week {col.replace('week_', '')}"

    wb = Workbook()
    C_HDR_DS  = "1F4E79"
    C_HDR_SUP = "375623"
    C_HDR_MOT = "C55A11"
    C_ALT     = "D9E1F2"
    C_TITULO  = "2F5597"
    C_RED     = "FF0000"

    def _hfnt(color="FFFFFF", size=11): return Font(bold=True, color=color, name="Calibri", size=size)
    def _bfnt(bold=False, size=10):     return Font(name="Calibri", size=size, bold=bold)

    # ── Aba TOP Ofensores ─────────────────────────────────────
    ws = wb.active
    ws.title = "TOP Ofensores"
    ws.sheet_view.showGridLines = False
    data_str = str(u.get("data_ref", ""))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
    t = ws.cell(1, 1, f"Reclamações de Fake Delivery  |  Referência: {data_str}")
    t.font = _hfnt(size=14); t.fill = PatternFill("solid", fgColor=C_TITULO)
    t.alignment = _CTR; ws.row_dimensions[1].height = 28

    # ── Tabela DS ─────────────────────────────────────────────
    semana_label = f"Qt Semana {semana_ref}" if semana_ref else "Qt Semana Atual"
    DS_HDR = ["DS", "SUPERVISOR", semana_label] + [_week_label(c) for c in week_cols_sta[:2]] + ["Qt Mês", "% Rate"]
    hf_ds = PatternFill("solid", fgColor=C_HDR_DS)
    for ci, cn in enumerate(DS_HDR, 1):
        c = ws.cell(3, ci, cn)
        c.fill = hf_ds; c.font = _hfnt(); c.alignment = _CTR; c.border = _BRD
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.row_dimensions[3].height = 36

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(DS_HDR))
    s2 = ws.cell(2, 1, "POR DS")
    s2.fill = hf_ds; s2.font = _hfnt(size=10); s2.alignment = _CTR

    row_ds = 4
    if not r_sta.empty:
        df_sta = r_sta.sort_values("dia_total", ascending=False).copy()
        for ri, (_, row) in enumerate(df_sta.iterrows(), row_ds):
            alt = ri % 2 == 0
            week_vals   = [int(row.get(wc, 0) or 0) for wc in week_cols_sta[:2]]
            mes_total   = int(row.get("mes_total", 0) or 0)
            dia_total   = int(row.get("dia_total", 0) or 0)
            supervisor_ds = str(row.get("supervisor", "") or "")
            vals = [row.get("station", ""), supervisor_ds, dia_total] + week_vals + [mes_total, ""]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(ri, ci, val)
                c.font = _bfnt(); c.border = _BRD; c.alignment = _CTR
                if alt: c.fill = PatternFill("solid", fgColor=C_ALT)
                if ci == 1: c.alignment = Alignment(horizontal="left", vertical="center")
                if ci in (3, 4, 5, 6): c.number_format = "#,##0"
        row_ds = row_ds + len(df_sta)

        total_dia = int(df_sta["dia_total"].sum())
        total_mes = int(df_sta["mes_total"].sum()) if "mes_total" in df_sta.columns else 0
        totals = ["TOTAL GERAL", "", total_dia] + [int(df_sta.get(wc, pd.Series([0])).sum()) for wc in week_cols_sta[:2]] + [total_mes, ""]
        for ci, val in enumerate(totals, 1):
            c = ws.cell(row_ds, ci, val)
            c.fill = PatternFill("solid", fgColor="BDD7EE")
            c.font = _hfnt(color="1F3864", size=10); c.border = _BRD; c.alignment = _CTR
            if ci == 1: c.alignment = Alignment(horizontal="left", vertical="center")
            if ci in (3, 4, 5, 6): c.number_format = "#,##0"
        row_ds += 1

    # ── Tabela Supervisor ─────────────────────────────────────
    SUP_START = len(DS_HDR) + 2
    SUP_HDR = ["SUPERVISOR", semana_label] + [_week_label(c) for c in week_cols_sup[:2]] + ["Qt Mês", "% Rate", "Performance"]
    hf_sup = PatternFill("solid", fgColor=C_HDR_SUP)

    ws.merge_cells(start_row=2, start_column=SUP_START, end_row=2, end_column=SUP_START + len(SUP_HDR) - 1)
    s2s = ws.cell(2, SUP_START, "POR SUPERVISOR")
    s2s.fill = hf_sup; s2s.font = _hfnt(size=10); s2s.alignment = _CTR

    for ci, cn in enumerate(SUP_HDR, SUP_START):
        c = ws.cell(3, ci, cn)
        c.fill = hf_sup; c.font = _hfnt(); c.alignment = _CTR; c.border = _BRD
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.row_dimensions[3].height = 36

    if not r_sup.empty:
        df_sup = r_sup.sort_values("dia_total", ascending=False).copy()
        prev_col = week_cols_sup[0] if week_cols_sup else None
        for ri, (_, row) in enumerate(df_sup.iterrows(), 4):
            alt = ri % 2 == 0
            dia  = int(row.get("dia_total", 0) or 0)
            prev = int(row.get(prev_col, 0) or 0) if prev_col else 0
            perf = "Melhor" if dia <= prev else "Piora" if dia > prev else "-"
            perf_color = "70AD47" if perf == "Melhor" else "FF0000" if perf == "Piora" else "808080"
            week_vals = [int(row.get(wc, 0) or 0) for wc in week_cols_sup[:2]]
            vals = [row.get("supervisor", ""), dia] + week_vals + [int(row.get("mes_total", 0) or 0), "", perf]
            for ci, val in enumerate(vals, SUP_START):
                c = ws.cell(ri, ci, val)
                c.font = _bfnt(); c.border = _BRD; c.alignment = _CTR
                if alt: c.fill = PatternFill("solid", fgColor=C_ALT)
                if ci == SUP_START: c.alignment = Alignment(horizontal="left", vertical="center")
                if ci in (SUP_START + 1, SUP_START + 2, SUP_START + 3, SUP_START + 4): c.number_format = "#,##0"
                if val in ("Melhor", "Piora"):
                    c.fill = PatternFill("solid", fgColor=perf_color)
                    c.font = _hfnt(color="FFFFFF", size=9)

    # ── TOP 5 Motoristas Ofensores ────────────────────────────
    if not top5.empty:
        MOT_START    = SUP_START
        mot_row_start = max(row_ds, 4 + (len(r_sup) if not r_sup.empty else 0)) + 2
        hf_mot = PatternFill("solid", fgColor=C_HDR_MOT)

        ws.merge_cells(start_row=mot_row_start, start_column=MOT_START,
                       end_row=mot_row_start, end_column=MOT_START + 4)
        tm = ws.cell(mot_row_start, MOT_START, "🏆 TOP Motoristas Ofensores")
        tm.fill = hf_mot; tm.font = _hfnt(); tm.alignment = _CTR
        ws.row_dimensions[mot_row_start].height = 24

        MOT_HDR = ["SUPERVISOR", "TOP Motorista Ofensor", "ID Motorista", "DS", f"Qt Reclamações Week {semana_ref}", "% do Total"]
        ncols_mot = len(MOT_HDR)
        ws.merge_cells(start_row=mot_row_start, start_column=MOT_START,
                       end_row=mot_row_start, end_column=MOT_START + ncols_mot - 1)
        for ci, cn in enumerate(MOT_HDR, MOT_START):
            c = ws.cell(mot_row_start + 1, ci, cn)
            c.fill = hf_mot; c.font = _hfnt(); c.alignment = _CTR; c.border = _BRD
            ws.column_dimensions[get_column_letter(ci)].width = 22

        total_dia_ref = int(r_sta["dia_total"].sum()) if not r_sta.empty else 1
        for ri, (_, row) in enumerate(top5.iterrows(), mot_row_start + 2):
            total_mot = int(row.get("total", 0) or 0)
            pct = total_mot / max(total_dia_ref, 1)
            vals = [
                str(row.get("supervisor", "") or ""),
                row.get("motorista", ""),
                str(row.get("id_motorista", "") or ""),
                str(row.get("ds", "") or ""),
                total_mot,
                pct,
            ]
            for ci, val in enumerate(vals, MOT_START):
                c = ws.cell(ri, ci, val)
                c.font = _bfnt(); c.border = _BRD; c.alignment = _CTR
                if ci in (MOT_START, MOT_START + 1): c.alignment = Alignment(horizontal="left", vertical="center")
                if ci == MOT_START + 4:
                    c.number_format = "#,##0"
                    c.font = Font(name="Calibri", size=11, bold=True, color=C_RED)
                if ci == MOT_START + 5:
                    c.number_format = "0.00%"
                    c.fill = PatternFill("solid", fgColor="FF0000") if pct > 0.005 else PatternFill("solid", fgColor="70AD47")
                    c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

    ws.freeze_panes = "A4"

    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Reclamacoes_{data_str}.xlsx"},
    )

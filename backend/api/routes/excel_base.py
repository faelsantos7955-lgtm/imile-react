"""
api/routes/excel_base.py — Estilos e utilitários compartilhados de geração de Excel
"""
import io
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Estilos globais ────────────────────────────────────────────
_HF   = PatternFill("solid", fgColor="1F3864")
_AF   = PatternFill("solid", fgColor="D9E1F2")
_RF   = PatternFill("solid", fgColor="FF0000")
_GF   = PatternFill("solid", fgColor="70AD47")
_PF   = PatternFill("solid", fgColor="5B9BD5")
_DSF  = PatternFill("solid", fgColor="BDD7EE")
_HFNT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_BFNT = Font(name="Arial", size=10)
_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LFT  = Alignment(horizontal="left", vertical="center")
_TH   = Side(style="thin", color="BFBFBF")
_BRD  = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)


def _titulo_aba(ws, titulo, n):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(n, 1))
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill = _HF
    c.alignment = _CTR
    ws.row_dimensions[1].height = 30


def _write_header(ws, cols, row=2):
    for ci, cn in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=cn)
        c.font = _HFNT; c.fill = _HF; c.alignment = _CTR; c.border = _BRD
    ws.row_dimensions[row].height = 26


def _write_data(ws, df, start_row=3, pct_cols=None):
    pct_cols = pct_cols or []
    for ri, row in enumerate(df.itertuples(index=False), start_row):
        alt = ri % 2 == 0
        for ci, val in enumerate(row, 1):
            try:
                if pd.isnull(val): val = None
            except: pass
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = _BFNT; c.alignment = _CTR; c.border = _BRD
            if alt: c.fill = _AF
            cn = df.columns[ci - 1]
            if cn in pct_cols and isinstance(val, (int, float)):
                c.number_format = "0.0%"
                if val < 0.5:
                    c.fill = _RF
                    c.font = Font(name="Arial", size=10, color="FFFFFF", bold=True)
            elif isinstance(val, (int, float)) and val > 100:
                c.number_format = "#,##0"


def _write_grouped(ws, df_ds, df_cid, start_row=3):
    """Escreve tabela agrupada DS → Cidades."""
    COLS = ["Scan Station", "recebido no DS", "em rota de entrega", "Total Geral",
            "Taxa de Expedicao", "Entregas", "Taxa de Entrega"]
    _write_header(ws, COLS, start_row - 1)

    city_index = {}
    if df_cid is not None and len(df_cid) > 0:
        for ds, grp in df_cid.groupby("scan_station", observed=True):
            city_index[ds] = grp.sort_values("recebido", ascending=False)

    DF   = Font(name="Arial", size=10, bold=True, color="1F3864")
    CITF = Font(name="Arial", size=9, italic=True, color="2E75B6")
    CF1  = PatternFill("solid", fgColor="F2F2F2")
    CF2  = PatternFill("solid", fgColor="FFFFFF")

    cur = start_row
    for _, dr in df_ds.iterrows():
        ds = dr["scan_station"]
        taxa_exp = float(dr.get("taxa_exp", 0) or 0)
        meta = float(dr.get("meta", 0.5) or 0.5)
        vals = [ds, int(dr.get("recebido", 0)), int(dr.get("expedido", 0)),
                int(dr.get("recebido", 0)), taxa_exp,
                int(dr.get("entregas", 0)), float(dr.get("taxa_ent", 0) or 0)]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=cur, column=ci, value=val)
            c.font = DF; c.fill = _DSF; c.border = _BRD
            c.alignment = _LFT if ci == 1 else _CTR
            if ci == 5:
                c.number_format = "0.0%"
                c.fill = _RF if taxa_exp < meta else _GF
                c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            elif ci == 7:
                c.number_format = "0.0%"; c.fill = _PF
                c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            elif ci in (2, 3, 4, 6):
                c.number_format = "#,##0"
        cur += 1

        city_rows = city_index.get(ds, pd.DataFrame())
        for i, (_, cr) in enumerate(city_rows.iterrows()):
            city = cr.get("destination_city", "")
            fill = CF1 if i % 2 == 0 else CF2
            taxa_c = float(cr.get("taxa_exp", 0) or 0)
            taxa_e = float(cr.get("taxa_ent", 0) or 0)
            vals2 = [f"   {city}", int(cr.get("recebido", 0)), int(cr.get("expedido", 0)),
                     int(cr.get("recebido", 0)), taxa_c,
                     int(cr.get("entregas", 0)), taxa_e]
            for ci, val in enumerate(vals2, 1):
                c = ws.cell(row=cur, column=ci, value=val)
                c.font = CITF; c.fill = fill; c.border = _BRD
                c.alignment = _LFT if ci == 1 else _CTR
                if ci in (5, 7): c.number_format = "0.0%"
                elif ci in (2, 3, 4, 6): c.number_format = "#,##0"
            ws.row_dimensions[cur].outline_level = 1
            ws.row_dimensions[cur].hidden = True
            cur += 1
        if len(city_rows) > 0:
            ws.sheet_view.showOutlineSymbols = True

    for col, w in zip([1, 2, 3, 4, 5, 6, 7], [26, 20, 22, 14, 18, 14, 16]):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = ws.cell(row=start_row, column=1)
    ws.sheet_properties.outlinePr.summaryBelow = False


def _auto_width(ws, extra=4):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        mx = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[cl].width = min(mx + extra, 35)


def _to_stream(wb):
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

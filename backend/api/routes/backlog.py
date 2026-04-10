"""
api/routes/backlog.py — Backlog SLA com persistência + filtro por cliente
"""
import logging
import io
from datetime import datetime

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text

from api.deps import get_db, get_current_user, require_admin, audit_log
from api.limiter import limiter
from api.upload_utils import validar_arquivo
import pandas as pd

logger = logging.getLogger(__name__)
router = APIRouter()

FAIXAS        = ['1-3', '3-5', '5-7', '7-10', '10-15', '15-20', 'Backlog >20']
FAIXAS_LABELS = ['1D≤X<3D', '3D≤X<5D', '5D≤X<7D', '7D≤X<10D', '10D≤X<15D', '15D≤X<20D', '≥20D']
DB_COLS       = ['f_1_3', 'f_3_5', 'f_5_7', 'f_7_10', 'f_10_15', 'f_15_20', 'f_20_mais']
FAIXA_7D      = {'7-10', '10-15', '15-20', 'Backlog >20'}


def _faixas_to_dict(row):
    return {f: row.get(col, 0) or 0 for f, col in zip(FAIXAS, DB_COLS)}


def _ler_excel(conteudo: bytes, supervisor_map: dict | None = None):
    buf = io.BytesIO(conteudo)
    try:
        xl = pd.ExcelFile(buf)
    except Exception:
        raise ValueError("Arquivo inválido ou corrompido.")

    abas = xl.sheet_names
    aba_det = next((a for a in abas if a.lower().replace(' ', '_') == 'backlog_details'), None)
    if aba_det is None:
        raise ValueError(f"Aba 'Backlog_Details' não encontrada. Abas: {abas}")

    aba_res = next(
        (a for a in abas if a.lower().replace('_', '').replace(' ', '') in ('resume',)),
        None,
    )

    buf.seek(0)
    df = pd.read_excel(buf, sheet_name=aba_det,
                       dtype={'waybillNo': str, 'range_backlog': str, 'process': str, 'actual_region': str})
    df.columns = df.columns.str.strip()

    cols_obrigatorias = ['waybillNo', 'range_backlog', 'process']
    faltando = [c for c in cols_obrigatorias if c not in df.columns]
    if faltando:
        raise ValueError(f"Colunas obrigatórias ausentes: {faltando}")

    df.rename(columns={'CARGOS.SUPERVISOR': 'supervisor', 'actual_region': 'regiao',
                       'lastScanSite': 'ds', 'clientName': 'cliente',
                       'stageStatus': 'estagio', 'lastScanStatus': 'motivo'}, inplace=True)

    df['ds']            = df['ds'].fillna('').str.strip().str.upper()
    df['regiao']        = df['regiao'].fillna('').str.strip()
    df['motivo']        = df['motivo'].fillna('Outros').str.strip()
    df['cliente']       = df['cliente'].fillna('Sem Cliente').str.strip()
    df['range_backlog'] = df['range_backlog'].fillna('1-3').str.strip()

    if 'supervisor' not in df.columns:
        df['supervisor'] = df['ds'].map(supervisor_map).fillna('Sem Supervisor') if supervisor_map else 'Sem Supervisor'
    else:
        df['supervisor'] = df['supervisor'].fillna('Sem Supervisor').str.strip().str.upper()

    df_res = pd.DataFrame()
    if aba_res:
        buf.seek(0)
        df_res = pd.read_excel(buf, sheet_name=aba_res)
        df_res.columns = df_res.columns.str.strip()
        df_res.rename(columns={'CARGOS.SUPERVISOR': 'supervisor', 'lastScanSite': 'ds',
                                'clientName': 'cliente', 'actual region': 'regiao'}, inplace=True)
        df_res['ds']     = df_res['ds'].fillna('').str.strip().str.upper() if 'ds' in df_res.columns else ''
        df_res['orders'] = pd.to_numeric(df_res.get('orders', 0), errors='coerce').fillna(0).astype(int)
        if 'supervisor' not in df_res.columns:
            df_res['supervisor'] = df_res['ds'].map(supervisor_map).fillna('Sem Supervisor') if supervisor_map else 'Sem Supervisor'
        else:
            df_res['supervisor'] = df_res['supervisor'].fillna('').str.strip().str.upper()

    return df, df_res


def _faixa_row(grp, orders):
    faixas   = {f: int((grp['range_backlog'] == f).sum()) for f in FAIXAS}
    total_7d = sum(faixas.get(f, 0) for f in FAIXA_7D)
    return {'orders': orders, 'backlog': len(grp),
            'pct_backlog': round(len(grp) / orders * 100, 1) if orders else 0,
            'faixas': faixas, 'total_7d': total_7d}


def _processar(df, df_res):
    dc  = df[df['process'].isin(['DC-LH', 'DC'])]
    dfs = df[df['process'] == 'DS']
    dc_res = df_res[df_res['process'].isin(['DC-LH', 'DC'])] if 'process' in df_res.columns else pd.DataFrame()
    ds_res = df_res[df_res['process'] == 'DS']               if 'process' in df_res.columns else df_res

    def orders(df_r, col, val):
        g = df_r[df_r[col] == val] if col in df_r.columns else pd.DataFrame()
        return int(g['orders'].sum()) if len(g) and 'orders' in g.columns else 0

    por_rdc = []
    for nome in sorted(dc['ds'].dropna().unique()):
        grp = dc[dc['ds'] == nome]
        row = _faixa_row(grp, orders(dc_res, 'ds', nome) or len(grp))
        row['nome']   = nome
        row['regiao'] = grp['regiao'].mode().iloc[0] if len(grp) else ''
        por_rdc.append(row)

    por_supervisor = []
    for nome in sorted(dfs['supervisor'].dropna().unique()):
        grp = dfs[dfs['supervisor'] == nome]
        row = _faixa_row(grp, orders(ds_res, 'supervisor', nome) or len(grp))
        row['nome'] = nome
        por_supervisor.append(row)

    por_ds = []
    for nome in sorted(dfs['ds'].dropna().unique()):
        grp = dfs[dfs['ds'] == nome]
        sup = grp['supervisor'].mode().iloc[0] if len(grp) else ''
        row = _faixa_row(grp, orders(ds_res, 'ds', nome) or len(grp))
        row['nome']       = nome
        row['supervisor'] = sup
        por_ds.append(row)
    por_ds.sort(key=lambda x: x['total_7d'], reverse=True)
    for i, r in enumerate(por_ds, 1):
        r['prioridade'] = i

    por_motivo = []
    for motivo in sorted(df['motivo'].dropna().unique()):
        grp = df[df['motivo'] == motivo]
        row = _faixa_row(grp, len(grp))
        row['nome'] = motivo
        por_motivo.append(row)
    por_motivo.sort(key=lambda x: x['backlog'], reverse=True)

    total = len(df)

    # Extrai data_ref do arquivo (coluna lastScanTime)
    data_ref = datetime.now().date().isoformat()
    if 'lastScanTime' in df.columns:
        try:
            col = df['lastScanTime']
            if pd.api.types.is_float_dtype(col) or pd.api.types.is_integer_dtype(col):
                datas = pd.to_datetime(col, unit='D', origin='1899-12-30', errors='coerce').dropna()
            else:
                datas = pd.to_datetime(col, errors='coerce').dropna()
            if not datas.empty:
                data_ref = datas.dt.date.max().isoformat()
        except Exception:
            pass

    kpis = {
        'total':       total,
        'na_ds':       int((df['estagio'] == 'Delivery').sum()),
        'em_transito': int((df['estagio'] == 'In Transit').sum()),
        'total_7d':    sum(r['total_7d'] for r in por_ds),
        'pct_7d':      round(sum(r['total_7d'] for r in por_ds) / total * 100, 1) if total else 0,
        'por_faixa':   {f: int((df['range_backlog'] == f).sum()) for f in FAIXAS},
        'data_ref':    data_ref,
    }
    return kpis, por_rdc, por_supervisor, por_ds, por_motivo


# ── Helpers SQL para filtro por cliente ──────────────────────

def _faixas_sql(prefix=""):
    """Gera SUM(CASE ...) para cada faixa de backlog."""
    faixa_map = list(zip(FAIXAS, DB_COLS))
    return ", ".join(
        f"SUM(CASE WHEN range_backlog = '{f}' THEN 1 ELSE 0 END) AS {col}"
        for f, col in faixa_map
    )


def _cliente_detalhes(db: Session, upload_id: int, cliente: str):
    """Agrega backlog_detalhes por cliente diretamente no banco."""
    faixas_expr = _faixas_sql()
    faixa_7d_expr = " + ".join(
        f"SUM(CASE WHEN range_backlog = '{f}' THEN 1 ELSE 0 END)"
        for f in FAIXA_7D
    )

    kpis_row = db.execute(text(f"""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN estagio = 'Delivery' THEN 1 ELSE 0 END) AS na_ds,
            SUM(CASE WHEN estagio = 'In Transit' THEN 1 ELSE 0 END) AS em_transito,
            {faixa_7d_expr} AS total_7d,
            {faixas_expr}
        FROM backlog_detalhes
        WHERE upload_id = :uid AND cliente = :cli
    """), {"uid": upload_id, "cli": cliente}).mappings().first()

    por_rdc = db.execute(text(f"""
        SELECT ds AS nome, regiao,
               COUNT(*) AS backlog, COUNT(*) AS orders,
               {faixas_expr},
               {faixa_7d_expr} AS total_7d
        FROM backlog_detalhes
        WHERE upload_id = :uid AND cliente = :cli AND process IN ('DC-LH','DC')
        GROUP BY ds, regiao ORDER BY backlog DESC
    """), {"uid": upload_id, "cli": cliente}).mappings().all()

    por_sup = db.execute(text(f"""
        SELECT supervisor AS nome,
               COUNT(*) AS backlog, COUNT(*) AS orders,
               {faixas_expr},
               {faixa_7d_expr} AS total_7d
        FROM backlog_detalhes
        WHERE upload_id = :uid AND cliente = :cli AND process = 'DS'
        GROUP BY supervisor ORDER BY backlog DESC
    """), {"uid": upload_id, "cli": cliente}).mappings().all()

    por_ds = db.execute(text(f"""
        SELECT ds AS nome, supervisor,
               COUNT(*) AS backlog, COUNT(*) AS orders,
               {faixas_expr},
               {faixa_7d_expr} AS total_7d
        FROM backlog_detalhes
        WHERE upload_id = :uid AND cliente = :cli AND process = 'DS'
        GROUP BY ds, supervisor ORDER BY total_7d DESC
    """), {"uid": upload_id, "cli": cliente}).mappings().all()

    por_mot = db.execute(text(f"""
        SELECT motivo AS nome,
               COUNT(*) AS backlog, COUNT(*) AS orders,
               {faixas_expr},
               {faixa_7d_expr} AS total_7d
        FROM backlog_detalhes
        WHERE upload_id = :uid AND cliente = :cli
        GROUP BY motivo ORDER BY backlog DESC
    """), {"uid": upload_id, "cli": cliente}).mappings().all()

    return dict(kpis_row), [dict(r) for r in por_rdc], [dict(r) for r in por_sup], \
           [dict(r) for r in por_ds], [dict(r) for r in por_mot]


# ── Endpoints ─────────────────────────────────────────────────

@router.get("/uploads")
def listar_uploads(user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    rows = db.execute(text(
        "SELECT * FROM backlog_uploads ORDER BY criado_em DESC LIMIT 30"
    )).mappings().all()
    return [dict(r) for r in rows]


@router.delete("/upload/{upload_id}")
def deletar_upload(upload_id: int, user: dict = Depends(require_admin), db: Session = Depends(get_db)):
    for tbl in ("backlog_detalhes", "backlog_por_cliente", "backlog_por_motivo",
                "backlog_por_ds", "backlog_por_supervisor", "backlog_por_rdc"):
        db.execute(text(f"DELETE FROM {tbl} WHERE upload_id = :uid"), {"uid": upload_id})
    db.execute(text("DELETE FROM backlog_uploads WHERE id = :uid"), {"uid": upload_id})
    db.commit()
    audit_log("upload_deletado", f"backlog_uploads:{upload_id}", {}, user)
    return {"ok": True}


@router.get("/clientes/{upload_id}")
def listar_clientes(
    upload_id: int,
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    total = db.execute(text(
        "SELECT COUNT(*) FROM backlog_por_cliente WHERE upload_id = :uid"
    ), {"uid": upload_id}).scalar()
    rows = db.execute(text(
        "SELECT * FROM backlog_por_cliente WHERE upload_id = :uid ORDER BY backlog DESC LIMIT :lim OFFSET :off"
    ), {"uid": upload_id, "lim": limit, "off": offset}).mappings().all()
    return {"total": total, "items": [dict(r) for r in rows], "limit": limit, "offset": offset}


@router.get("/upload/{upload_id}")
def detalhe_upload(
    upload_id: int,
    cliente: str = Query(None),
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if cliente:
        kpis, por_rdc, por_sup, por_ds, por_mot = _cliente_detalhes(db, upload_id, cliente)

        up = db.execute(text(
            "SELECT data_ref FROM backlog_uploads WHERE id = :uid"
        ), {"uid": upload_id}).mappings().first()
        data_ref = str(up["data_ref"]) if up else ""

        def enrich(rows):
            for r in rows:
                ords = r.get("orders") or r.get("backlog") or 1
                r["orders"]      = ords
                r["pct_backlog"] = round(r["backlog"] / ords * 100, 1) if ords else 0
                r["faixas"]      = {f: r.get(col, 0) or 0 for f, col in zip(FAIXAS, DB_COLS)}
            return rows

        por_ds_enriched = enrich(por_ds)
        for i, r in enumerate(por_ds_enriched, 1):
            r["prioridade"] = i

        total = kpis.get("total") or 1
        kpis["data_ref"]    = data_ref
        kpis["pct_7d"]      = round((kpis.get("total_7d") or 0) / total * 100, 1)
        kpis["por_faixa"]   = {f: kpis.get(col, 0) or 0 for f, col in zip(FAIXAS, DB_COLS)}

        return {
            "kpis":          kpis,
            "por_rdc":       enrich(por_rdc),
            "por_supervisor": enrich(por_sup),
            "por_ds":        por_ds_enriched,
            "por_motivo":    enrich(por_mot),
        }

    up = db.execute(text("SELECT * FROM backlog_uploads WHERE id = :uid"), {"uid": upload_id}).mappings().first()
    if not up:
        raise HTTPException(404, "Não encontrado")
    u = dict(up)

    rdc  = [dict(r) for r in db.execute(text("SELECT * FROM backlog_por_rdc WHERE upload_id = :uid"), {"uid": upload_id}).mappings().all()]
    sups = [dict(r) for r in db.execute(text("SELECT * FROM backlog_por_supervisor WHERE upload_id = :uid"), {"uid": upload_id}).mappings().all()]
    dss  = [dict(r) for r in db.execute(text("SELECT * FROM backlog_por_ds WHERE upload_id = :uid ORDER BY prioridade"), {"uid": upload_id}).mappings().all()]
    mot  = [dict(r) for r in db.execute(text("SELECT * FROM backlog_por_motivo WHERE upload_id = :uid ORDER BY backlog DESC"), {"uid": upload_id}).mappings().all()]

    fmt   = lambda rows: [{**r, "faixas": _faixas_to_dict(r)} for r in rows]
    total = u["total"] or 1

    return {
        "kpis": {
            "total":       u["total"],
            "na_ds":       u["na_ds"],
            "em_transito": u["em_transito"],
            "total_7d":    u["total_7d"],
            "pct_7d":      round(u["total_7d"] / total * 100, 1),
            "por_faixa":   {f: sum(r.get(col, 0) or 0 for r in dss) for f, col in zip(FAIXAS, DB_COLS)},
            "data_ref":    u["data_ref"],
        },
        "por_rdc":        fmt(rdc),
        "por_supervisor": fmt(sups),
        "por_ds":         fmt(dss),
        "por_motivo":     fmt(mot),
    }


@router.post("/processar")
@limiter.limit("5/minute")
async def processar_backlog(
    request: Request,
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    conteudo = await validar_arquivo(file)
    try:
        df, df_res = _ler_excel(conteudo)
    except Exception as e:
        raise HTTPException(400, f"Erro ao ler arquivo: {e}")

    kpis, por_rdc, por_supervisor, por_ds, por_motivo = _processar(df, df_res)

    # Remove upload anterior da mesma data_ref
    existing = db.execute(text(
        "SELECT id FROM backlog_uploads WHERE data_ref = :dr"
    ), {"dr": kpis["data_ref"]}).mappings().first()
    if existing:
        old_id = existing["id"]
        for tbl in ("backlog_detalhes", "backlog_por_ds", "backlog_por_supervisor",
                    "backlog_por_rdc", "backlog_por_motivo", "backlog_por_cliente"):
            db.execute(text(f"DELETE FROM {tbl} WHERE upload_id = :uid"), {"uid": old_id})
        db.execute(text("DELETE FROM backlog_uploads WHERE id = :uid"), {"uid": old_id})

    row = db.execute(text("""
        INSERT INTO backlog_uploads (data_ref, criado_por, total, total_7d, na_ds, em_transito)
        VALUES (:data_ref, :criado_por, :total, :total_7d, :na_ds, :em_transito)
        RETURNING id
    """), {
        "data_ref":    kpis["data_ref"],
        "criado_por":  user["email"],
        "total":       kpis["total"],
        "total_7d":    kpis["total_7d"],
        "na_ds":       kpis["na_ds"],
        "em_transito": kpis["em_transito"],
    }).mappings().first()
    uid = row["id"]

    def to_db(rows, extras=[]):
        result = []
        for r in rows:
            rec = {"upload_id": uid, "nome": r["nome"], "orders": r["orders"],
                   "backlog": r["backlog"], "pct_backlog": r["pct_backlog"], "total_7d": r["total_7d"]}
            for f, col in zip(FAIXAS, DB_COLS):
                rec[col] = r["faixas"].get(f, 0)
            for k in extras:
                rec[k] = r.get(k, "")
            result.append(rec)
        return result

    def bulk_insert(tbl, rows, cols):
        if not rows:
            return
        placeholders = ", ".join(f":{c}" for c in cols)
        col_names    = ", ".join(cols)
        db.execute(text(f"INSERT INTO {tbl} ({col_names}) VALUES ({placeholders})"), rows)

    rdc_cols = ["upload_id", "nome", "orders", "backlog", "pct_backlog", "total_7d", "regiao"] + DB_COLS
    sup_cols = ["upload_id", "nome", "orders", "backlog", "pct_backlog", "total_7d"] + DB_COLS
    ds_cols  = ["upload_id", "nome", "orders", "backlog", "pct_backlog", "total_7d", "supervisor", "prioridade"] + DB_COLS
    mot_cols = ["upload_id", "nome", "orders", "backlog", "pct_backlog", "total_7d"] + DB_COLS

    bulk_insert("backlog_por_rdc",        to_db(por_rdc, ["regiao"]),                         rdc_cols)
    bulk_insert("backlog_por_supervisor", to_db(por_supervisor),                               sup_cols)
    bulk_insert("backlog_por_ds",         to_db(por_ds, ["supervisor", "prioridade"]),         ds_cols)
    bulk_insert("backlog_por_motivo",     to_db(por_motivo),                                   mot_cols)

    cols = ["waybillNo", "cliente", "supervisor", "ds", "process", "range_backlog", "motivo", "estagio", "regiao"]
    for col in cols:
        if col not in df.columns:
            df[col] = ""
    detalhes = [
        {
            "upload_id":     uid,
            "waybill":       str(r["waybillNo"]),
            "cliente":       str(r["cliente"]),
            "supervisor":    str(r["supervisor"]),
            "ds":            str(r["ds"]),
            "process":       str(r["process"]),
            "range_backlog": str(r["range_backlog"]),
            "motivo":        str(r["motivo"]),
            "estagio":       str(r["estagio"]),
            "regiao":        str(r["regiao"]),
        }
        for r in df[cols].to_dict("records")
    ]
    det_cols = ["upload_id", "waybill", "cliente", "supervisor", "ds", "process", "range_backlog", "motivo", "estagio", "regiao"]
    for i in range(0, len(detalhes), 1000):
        bulk_insert("backlog_detalhes", detalhes[i:i+1000], det_cols)

    db.commit()
    return {"upload_id": uid, "kpis": kpis,
            "por_rdc": por_rdc, "por_supervisor": por_supervisor,
            "por_ds": por_ds, "por_motivo": por_motivo}


@router.post("/excel/{upload_id}")
@limiter.limit("20/minute")
def excel_backlog(
    request: Request,
    upload_id: int,
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    up = db.execute(text("SELECT * FROM backlog_uploads WHERE id = :uid"), {"uid": upload_id}).mappings().first()
    if not up:
        raise HTTPException(404, "Não encontrado")
    u = dict(up)

    rdc  = [dict(r) for r in db.execute(text("SELECT * FROM backlog_por_rdc WHERE upload_id = :uid"), {"uid": upload_id}).mappings().all()]
    sups = [dict(r) for r in db.execute(text("SELECT * FROM backlog_por_supervisor WHERE upload_id = :uid"), {"uid": upload_id}).mappings().all()]
    dss  = [dict(r) for r in db.execute(text("SELECT * FROM backlog_por_ds WHERE upload_id = :uid ORDER BY prioridade"), {"uid": upload_id}).mappings().all()]
    mot  = [dict(r) for r in db.execute(text("SELECT * FROM backlog_por_motivo WHERE upload_id = :uid ORDER BY backlog DESC"), {"uid": upload_id}).mappings().all()]

    CORES_FAIXA = {"1-3": "92D050", "3-5": "FFFF00", "5-7": "FFC000", "7-10": "EF4444",
                   "10-15": "DC2626", "15-20": "B91C1C", "Backlog >20": "7F1D1D"}
    HDR  = PatternFill("solid", fgColor="1F3864")
    AZUL = PatternFill("solid", fgColor="2E75B6")
    VERD = PatternFill("solid", fgColor="375623")
    LROX = PatternFill("solid", fgColor="7030A0")
    ALT  = PatternFill("solid", fgColor="D9E1F2")
    BRD  = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LFT  = Alignment(horizontal="left", vertical="center")
    HFNT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    BFNT = Font(name="Calibri", size=10)

    wb = Workbook()
    ws = wb.active
    ws.title = "BACKLOG 超时未完结"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("C1:P1")
    t = ws.cell(1, 3, f'BACKLOG  —  {u["data_ref"]}')
    t.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    t.fill = HDR; t.alignment = CTR
    ws.row_dimensions[1].height = 32

    def write_section(cur, titulo, rows, fill, show_sup=False, show_regiao=False):
        ws.merge_cells(f"C{cur}:P{cur}")
        h = ws.cell(cur, 3, titulo)
        h.fill = fill; h.font = HFNT; h.alignment = CTR
        ws.row_dimensions[cur].height = 22
        cur += 1
        hdr = []
        if show_regiao: hdr.append("Região")
        if show_sup:    hdr.append("Supervisor")
        hdr += ["Nome", "Orders", "Backlog", "% Backlog"] + FAIXAS_LABELS + [">7D"]
        if show_sup:    hdr.append("Prioridade")
        for ci, txt in enumerate(hdr, 3):
            c = ws.cell(cur, ci, txt)
            c.fill = fill; c.font = HFNT; c.alignment = CTR; c.border = BRD
        ws.row_dimensions[cur].height = 36
        cur += 1
        for i, row in enumerate(rows):
            rf = ALT if i % 2 == 0 else None
            vals = []
            if show_regiao: vals.append(row.get("regiao", ""))
            if show_sup:    vals.append(row.get("supervisor", ""))
            vals += [row["nome"], row["orders"], row["backlog"], round(row["pct_backlog"] / 100, 4)]
            for col in DB_COLS: vals.append(row.get(col, 0) or 0)
            vals.append(row["total_7d"])
            if show_sup: vals.append(row.get("prioridade", ""))
            offset = 3 + (1 if show_regiao else 0) + (1 if show_sup else 0)
            for ci, val in enumerate(vals, 3):
                c = ws.cell(cur, ci, val)
                c.font = BFNT; c.border = BRD; c.alignment = CTR
                if rf: c.fill = rf
                if ci in (3, 4): c.alignment = LFT
                if ci == offset + 2: c.number_format = "0.0%"
                fo = ci - (offset + 3)
                if 0 <= fo < len(FAIXAS):
                    cor = CORES_FAIXA.get(FAIXAS[fo], "FFFFFF")
                    if isinstance(val, (int, float)) and val > 0:
                        c.fill = PatternFill("solid", fgColor=cor)
                        c.font = Font(name="Calibri", size=10, bold=True,
                                      color="FFFFFF" if FAIXAS[fo] not in ("1-3", "3-5") else "000000")
            cur += 1
        return cur + 1

    cur = 3
    cur = write_section(cur, "LH — Por RDC",           rdc,  AZUL, show_regiao=True)
    cur = write_section(cur, "DS — Por Supervisor",     sups, VERD)
    cur = write_section(cur, "DS — Detalhado por Base", dss,  HDR,  show_sup=True)
    cur = write_section(cur, "DS — Por Motivo",         mot,  LROX)

    for col, w in [(3, 20), (4, 14), (5, 12), (6, 12), (7, 10)] + [(8 + i, 9) for i in range(9)]:
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "C4"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=Backlog_SLA_{u["data_ref"]}.xlsx'})

"""
api/routes/excel.py — Download de Excel formatado
Dashboard (DS + cidades agrupadas + regiões), Reclamações (TOP Ofensores),
Triagem (DS + supervisor), Histórico (período consolidado)
"""
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from api.deps import get_supabase, get_current_user
from api.limiter import limiter
import pandas as pd
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()

# ── Estilos ───────────────────────────────────────────────────
_HF   = PatternFill("solid", fgColor="1F3864")
_AF   = PatternFill("solid", fgColor="D9E1F2")
_RF   = PatternFill("solid", fgColor="FF0000")
_GF   = PatternFill("solid", fgColor="70AD47")
_PF   = PatternFill("solid", fgColor="5B9BD5")
_DSF  = PatternFill("solid", fgColor="BDD7EE")
_HFNT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_BFNT = Font(name="Arial", size=10)
_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LFT  = Alignment(horizontal="left", vertical="center")
_TH   = Side(style="thin", color="BFBFBF")
_BRD  = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)


def _titulo_aba(ws, titulo, n):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(n, 1))
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill = _HF
    c.alignment = _CTR
    ws.row_dimensions[1].height = 30


def _write_header(ws, cols, row=2):
    for ci, cn in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=cn)
        c.font = _HFNT; c.fill = _HF; c.alignment = _CTR; c.border = _BRD
    ws.row_dimensions[row].height = 26


def _write_data(ws, df, start_row=3, pct_cols=None):
    pct_cols = pct_cols or []
    for ri, row in enumerate(df.itertuples(index=False), start_row):
        alt = ri % 2 == 0
        for ci, val in enumerate(row, 1):
            try:
                if pd.isnull(val): val = None
            except: pass
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = _BFNT; c.alignment = _CTR; c.border = _BRD
            if alt: c.fill = _AF
            cn = df.columns[ci - 1]
            if cn in pct_cols and isinstance(val, (int, float)):
                c.number_format = "0.0%"
                if val < 0.5:
                    c.fill = _RF
                    c.font = Font(name="Arial", size=10, color="FFFFFF", bold=True)
            elif isinstance(val, (int, float)) and val > 100:
                c.number_format = "#,##0"


def _write_grouped(ws, df_ds, df_cid, start_row=3):
    """Escreve tabela agrupada DS → Cidades (mesmo formato do reference)."""
    COLS = ["Scan Station", "recebido no DS", "em rota de entrega", "Total Geral",
            "Taxa de Expedicao", "Entregas", "Taxa de Entrega"]
    _write_header(ws, COLS, start_row - 1)

    city_index = {}
    if df_cid is not None and len(df_cid) > 0:
        for ds, grp in df_cid.groupby("scan_station", observed=True):
            city_index[ds] = grp.sort_values("recebido", ascending=False)

    DF = Font(name="Arial", size=10, bold=True, color="1F3864")
    CITF = Font(name="Arial", size=9, italic=True, color="2E75B6")
    CF1 = PatternFill("solid", fgColor="F2F2F2")
    CF2 = PatternFill("solid", fgColor="FFFFFF")

    cur = start_row
    for _, dr in df_ds.iterrows():
        ds = dr["scan_station"]
        taxa_exp = float(dr.get("taxa_exp", 0) or 0)
        meta = float(dr.get("meta", 0.5) or 0.5)
        vals = [ds, int(dr.get("recebido", 0)), int(dr.get("expedido", 0)),
                int(dr.get("recebido", 0)), taxa_exp,
                int(dr.get("entregas", 0)), float(dr.get("taxa_ent", 0) or 0)]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=cur, column=ci, value=val)
            c.font = DF; c.fill = _DSF; c.border = _BRD
            c.alignment = _LFT if ci == 1 else _CTR
            if ci == 5:
                c.number_format = "0.0%"
                c.fill = _RF if taxa_exp < meta else _GF
                c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            elif ci == 7:
                c.number_format = "0.0%"; c.fill = _PF
                c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            elif ci in (2, 3, 4, 6):
                c.number_format = "#,##0"
        cur += 1

        # Linhas de cidade — agrupadas e colapsadas por padrão
        city_rows = city_index.get(ds, pd.DataFrame())
        city_start = cur
        for i, (_, cr) in enumerate(city_rows.iterrows()):
            city = cr.get("destination_city", "")
            fill = CF1 if i % 2 == 0 else CF2
            taxa_c = float(cr.get("taxa_exp", 0) or 0)
            taxa_e = float(cr.get("taxa_ent", 0) or 0)
            vals2 = [f"   {city}", int(cr.get("recebido", 0)), int(cr.get("expedido", 0)),
                     int(cr.get("recebido", 0)), taxa_c,
                     int(cr.get("entregas", 0)), taxa_e]
            for ci, val in enumerate(vals2, 1):
                c = ws.cell(row=cur, column=ci, value=val)
                c.font = CITF; c.fill = fill; c.border = _BRD
                c.alignment = _LFT if ci == 1 else _CTR
                if ci in (5, 7): c.number_format = "0.0%"
                elif ci in (2, 3, 4, 6): c.number_format = "#,##0"
            # Agrupa e colapsa a linha de cidade
            ws.row_dimensions[cur].outline_level = 1
            ws.row_dimensions[cur].hidden = True
            cur += 1
        # Configura o agrupamento para colapsar acima (botão + na linha da DS)
        if len(city_rows) > 0:
            ws.sheet_view.showOutlineSymbols = True

    # Column widths
    for col, w in zip([1, 2, 3, 4, 5, 6, 7], [26, 20, 22, 14, 18, 14, 16]):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = ws.cell(row=start_row, column=1)
    # Botão + aparece na linha da DS (summary acima das cidades)
    ws.sheet_properties.outlinePr.summaryBelow = False


def _auto_width(ws, extra=4):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        mx = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[cl].width = min(mx + extra, 35)


def _to_stream(wb):
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════
#  DASHBOARD EXCEL (DS + cidades agrupadas + regiões)
# ══════════════════════════════════════════════════════════════
@router.get("/dashboard/{data_ref}")
@limiter.limit("20/minute")
def excel_dashboard(request: Request, data_ref: str, user: dict = Depends(get_current_user)):
    sb = get_supabase()

    q = sb.table("expedicao_diaria").select("*").eq("data_ref", data_ref)
    if user["bases"]: q = q.in_("scan_station", user["bases"])
    dia = pd.DataFrame(q.execute().data or [])
    if dia.empty: raise HTTPException(404, "Sem dados")

    qc = sb.table("expedicao_cidades").select("*").eq("data_ref", data_ref)
    if user["bases"]: qc = qc.in_("scan_station", user["bases"])
    cid = pd.DataFrame(qc.execute().data or [])

    dia = dia.sort_values("recebido", ascending=False)
    wb = Workbook()

    # ── Aba Consolidado Geral (DS + cidades) ──────────────────
    ws = wb.active; ws.title = "Consolidado_Geral"
    _titulo_aba(ws, f"Consolidado Geral — {data_ref}", 7)
    _write_grouped(ws, dia, cid, start_row=3)

    # ── Abas por região (DS + cidades) ────────────────────────
    for regiao, label in [("capital", "Capital"), ("metropolitan", "Metropolitan"), ("countryside", "Countryside")]:
        df_r = dia[dia["region"].str.lower() == regiao]
        if df_r.empty: continue
        cid_r = cid[cid["scan_station"].isin(df_r["scan_station"])] if not cid.empty else pd.DataFrame()
        ws_r = wb.create_sheet(label)
        _titulo_aba(ws_r, label, 7)
        _write_grouped(ws_r, df_r, cid_r, start_row=3)

    # ── Aba Resumo Dinâmico (regiões lado a lado) ─────────────
    ws_res = wb.create_sheet("Resumo_Dinamico")
    ws_res.sheet_view.showGridLines = False

    REGIOES = [("capital", "Capital"), ("metropolitan", "Metropolitan"), ("countryside", "Countryside")]
    COLS_RES = ["DS", "recebido no DS", "em rota de entrega", "Total Geral", "Taxa de Expedicao"]

    # Título geral
    total_cols = len(COLS_RES) * 3 + 2
    ws_res.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws_res.cell(1, 1, f"Resumo Consolidado — {data_ref}")
    t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill = _HF; t.alignment = _CTR; ws_res.row_dimensions[1].height = 30

    col_start = 1
    for regiao, label in REGIOES:
        df_r = dia[dia["region"].str.lower() == regiao].copy()
        if df_r.empty: continue

        # Resumo da região (totais)
        rec = int(df_r["recebido"].sum())
        exp = int(df_r["expedido"].sum())
        ent = int(df_r["entregas"].sum())
        taxa = round(exp/rec, 4) if rec else 0

        # Header da região
        ws_res.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_start + len(COLS_RES) - 1)
        h = ws_res.cell(2, col_start, f"{label}   |   Rec: {rec:,}   Exp: {exp:,}   Taxa: {taxa:.1%}")
        h.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        h.fill = _HF; h.alignment = _CTR; ws_res.row_dimensions[2].height = 24

        # Cabeçalhos das colunas
        for ci, cn in enumerate(COLS_RES, col_start):
            c = ws_res.cell(3, ci, cn)
            c.font = _HFNT; c.fill = _HF; c.alignment = _CTR; c.border = _BRD
        ws_res.row_dimensions[3].height = 22

        # Dados
        df_r = df_r.sort_values("recebido", ascending=False)
        for ri, (_, r) in enumerate(df_r.iterrows(), 4):
            taxa_ds = float(r.get("taxa_exp", 0) or 0)
            meta_ds = float(r.get("meta", 0.5) or 0.5)
            alt = ri % 2 == 0
            vals = [r["scan_station"], int(r.get("recebido",0)), int(r.get("expedido",0)),
                    int(r.get("recebido",0)), taxa_ds]
            for ci, val in enumerate(vals, col_start):
                c = ws_res.cell(ri, ci, val)
                c.font = _BFNT; c.border = _BRD; c.alignment = _CTR
                if alt: c.fill = _AF
                if ci == col_start: c.alignment = Alignment(horizontal="left", vertical="center")
                if ci == col_start + 4:
                    c.number_format = "0.0%"
                    c.fill = _RF if taxa_ds < meta_ds else _GF
                    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                elif ci in (col_start+1, col_start+2, col_start+3):
                    c.number_format = "#,##0"

        # Linha de total da região
        last_r = 4 + len(df_r)
        totals = ["Total Geral", rec, exp, rec, taxa]
        for ci, val in enumerate(totals, col_start):
            c = ws_res.cell(last_r, ci, val)
            c.fill = PatternFill("solid", fgColor="BDD7EE")
            c.font = Font(name="Arial", bold=True, size=10, color="1F3864")
            c.border = _BRD; c.alignment = _CTR
            if ci == col_start: c.alignment = Alignment(horizontal="left", vertical="center")
            if ci == col_start + 4: c.number_format = "0.0%"
            elif ci in (col_start+1, col_start+2, col_start+3): c.number_format = "#,##0"

        # Larguras das colunas
        for ci2, w in enumerate([22, 16, 18, 14, 16], col_start):
            ws_res.column_dimensions[get_column_letter(ci2)].width = w

        col_start += len(COLS_RES) + 1  # +1 para espaço entre regiões

    ws_res.freeze_panes = "A4"

    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Dashboard_{data_ref}.xlsx"},
    )


# ══════════════════════════════════════════════════════════════
#  RECLAMAÇÕES EXCEL (formato TOP Ofensores igual imagem 5)
# ══════════════════════════════════════════════════════════════
@router.get("/reclamacoes/{upload_id}")
@limiter.limit("20/minute")
def excel_reclamacoes(request: Request, upload_id: int, user: dict = Depends(get_current_user)):
    sb = get_supabase()

    upload = sb.table("reclamacoes_uploads").select("*").eq("id", upload_id).execute()
    if not upload.data: raise HTTPException(404, "Upload não encontrado")
    u = upload.data[0]

    r_sup  = pd.DataFrame(sb.table("reclamacoes_por_supervisor").select("*").eq("upload_id", upload_id).execute().data or [])
    r_sta  = pd.DataFrame(sb.table("reclamacoes_por_station").select("*").eq("upload_id", upload_id).execute().data or [])
    top_all = pd.DataFrame(sb.table("reclamacoes_top5").select("*").eq("upload_id", upload_id).order("total", desc=True).execute().data or [])

    # Filtra inativos e pega top 5
    inativos_res = sb.table("motoristas_status").select("id_motorista").eq("ativo", False).execute()
    inativos = {r["id_motorista"] for r in (inativos_res.data or [])}
    top5 = top_all[~top_all["motorista"].isin(inativos)].head(5) if not top_all.empty else top_all

    # Detecta semanas disponíveis nas colunas (week_XX)
    semana_ref = u.get("semana_ref", 0)
    week_cols_sta = sorted([c for c in r_sta.columns if c.startswith("week_")], reverse=True) if not r_sta.empty else []
    week_cols_sup = sorted([c for c in r_sup.columns if c.startswith("week_")], reverse=True) if not r_sup.empty else []

    def _week_label(col):
        return f"Week {col.replace('week_', '')}"

    wb = Workbook()
    C_HDR_DS  = "1F4E79"
    C_HDR_SUP = "375623"
    C_HDR_MOT = "C55A11"
    C_ALT     = "D9E1F2"
    C_TITULO  = "2F5597"
    C_RED     = "FF0000"

    def _hfnt(color="FFFFFF", size=11):
        return Font(bold=True, color=color, name="Calibri", size=size)
    def _bfnt(bold=False, size=10):
        return Font(name="Calibri", size=size, bold=bold)

    # ── Aba TOP Ofensores ─────────────────────────────────────
    ws = wb.active
    ws.title = "TOP Ofensores"
    ws.sheet_view.showGridLines = False
    data_str = str(u.get("data_ref", ""))

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
    t = ws.cell(1, 1, f"Reclamações de Fake Delivery  |  Referência: {data_str}")
    t.font = _hfnt(size=14); t.fill = PatternFill("solid", fgColor=C_TITULO)
    t.alignment = _CTR; ws.row_dimensions[1].height = 28

    # ── Tabela DS (esquerda, cols 1+) ─────────────────────────
    semana_label = f"Qt Semana {semana_ref}" if semana_ref else "Qt Semana Atual"
    DS_HDR = ["DS", "SUPERVISOR", semana_label] + [_week_label(c) for c in week_cols_sta[:2]] + ["Qt Mês", "% Rate"]
    hf_ds = PatternFill("solid", fgColor=C_HDR_DS)
    for ci, cn in enumerate(DS_HDR, 1):
        c = ws.cell(3, ci, cn)
        c.fill = hf_ds; c.font = _hfnt(); c.alignment = _CTR; c.border = _BRD
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.row_dimensions[3].height = 36

    # Subtítulo em linha 2
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(DS_HDR))
    s2 = ws.cell(2, 1, "POR DS")
    s2.fill = hf_ds; s2.font = _hfnt(size=10); s2.alignment = _CTR

    row_ds = 4
    if not r_sta.empty:
        df_sta = r_sta.sort_values("dia_total", ascending=False).copy()
        # Tenta buscar supervisor predominante por DS do top_all
        ds_sup_map = {}
        if not top_all.empty and "ds" in top_all.columns and "motorista" in top_all.columns:
            pass  # sem supervisor no top5, deixa vazio
        if not r_sup.empty:
            # Heurística: associa supervisor com mais reclamações a cada DS
            pass

        for ri, (_, row) in enumerate(df_sta.iterrows(), row_ds):
            alt = ri % 2 == 0
            week_vals = [int(row.get(wc, 0) or 0) for wc in week_cols_sta[:2]]
            mes_total = int(row.get("mes_total", 0) or 0)
            dia_total = int(row.get("dia_total", 0) or 0)
            supervisor_ds = str(row.get("supervisor", "") or "")
            # % Rate = reclamações semana / entregas mês (se disponível)
            rate = ""
            vals = [row.get("station", ""), supervisor_ds, dia_total] + week_vals + [mes_total, rate]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(ri, ci, val)
                c.font = _bfnt(); c.border = _BRD; c.alignment = _CTR
                if alt: c.fill = PatternFill("solid", fgColor=C_ALT)
                if ci == 1: c.alignment = Alignment(horizontal="left", vertical="center")
                if ci in (3, 4, 5, 6): c.number_format = "#,##0"
        row_ds = row_ds + len(df_sta)

        # Linha total
        total_dia = int(df_sta["dia_total"].sum())
        total_mes = int(df_sta["mes_total"].sum()) if "mes_total" in df_sta.columns else 0
        totals = ["TOTAL GERAL", "", total_dia] + [int(df_sta.get(wc, pd.Series([0])).sum()) for wc in week_cols_sta[:2]] + [total_mes, ""]
        for ci, val in enumerate(totals, 1):
            c = ws.cell(row_ds, ci, val)
            c.fill = PatternFill("solid", fgColor="BDD7EE")
            c.font = _hfnt(color="1F3864", size=10); c.border = _BRD; c.alignment = _CTR
            if ci == 1: c.alignment = Alignment(horizontal="left", vertical="center")
            if ci in (3, 4, 5, 6): c.number_format = "#,##0"
        row_ds += 1

    # ── Tabela Supervisor (direita, cols 8+) ───────────────────
    SUP_START = len(DS_HDR) + 2
    SUP_HDR = ["SUPERVISOR", semana_label] + [_week_label(c) for c in week_cols_sup[:2]] + ["Qt Mês", "% Rate", "Performance"]
    hf_sup = PatternFill("solid", fgColor=C_HDR_SUP)

    ws.merge_cells(start_row=2, start_column=SUP_START, end_row=2, end_column=SUP_START + len(SUP_HDR) - 1)
    s2s = ws.cell(2, SUP_START, "POR SUPERVISOR")
    s2s.fill = hf_sup; s2s.font = _hfnt(size=10); s2s.alignment = _CTR

    for ci, cn in enumerate(SUP_HDR, SUP_START):
        c = ws.cell(3, ci, cn)
        c.fill = hf_sup; c.font = _hfnt(); c.alignment = _CTR; c.border = _BRD
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.row_dimensions[3].height = 36

    if not r_sup.empty:
        df_sup = r_sup.sort_values("dia_total", ascending=False).copy()
        prev_col = week_cols_sup[0] if week_cols_sup else None
        for ri, (_, row) in enumerate(df_sup.iterrows(), 4):
            alt = ri % 2 == 0
            dia = int(row.get("dia_total", 0) or 0)
            prev = int(row.get(prev_col, 0) or 0) if prev_col else 0
            perf = "Melhor" if dia <= prev else "Piora" if dia > prev else "-"
            perf_color = "70AD47" if perf == "Melhor" else "FF0000" if perf == "Piora" else "808080"
            week_vals = [int(row.get(wc, 0) or 0) for wc in week_cols_sup[:2]]
            vals = [row.get("supervisor", ""), dia] + week_vals + [int(row.get("mes_total", 0) or 0), "", perf]
            for ci, val in enumerate(vals, SUP_START):
                c = ws.cell(ri, ci, val)
                c.font = _bfnt(); c.border = _BRD; c.alignment = _CTR
                if alt: c.fill = PatternFill("solid", fgColor=C_ALT)
                if ci == SUP_START: c.alignment = Alignment(horizontal="left", vertical="center")
                if ci in (SUP_START + 1, SUP_START + 2, SUP_START + 3, SUP_START + 4): c.number_format = "#,##0"
                if val in ("Melhor", "Piora"):
                    c.fill = PatternFill("solid", fgColor=perf_color)
                    c.font = _hfnt(color="FFFFFF", size=9)

    # ── TOP 5 Motoristas Ofensores ────────────────────────────
    if not top5.empty:
        MOT_START = SUP_START
        mot_row_start = max(row_ds, 4 + (len(r_sup) if not r_sup.empty else 0)) + 2
        hf_mot = PatternFill("solid", fgColor=C_HDR_MOT)

        ws.merge_cells(start_row=mot_row_start, start_column=MOT_START,
                       end_row=mot_row_start, end_column=MOT_START + 4)
        tm = ws.cell(mot_row_start, MOT_START, "🏆 TOP 5 Motoristas Ofensores")
        tm.fill = hf_mot; tm.font = _hfnt(); tm.alignment = _CTR
        ws.row_dimensions[mot_row_start].height = 24

        MOT_HDR = ["SUPERVISOR", "TOP Motorista Ofensor", "ID Motorista", "DS", f"Qt Reclamações Week {semana_ref}", "% do Total"]
        ncols_mot = len(MOT_HDR)
        ws.merge_cells(start_row=mot_row_start, start_column=MOT_START,
                       end_row=mot_row_start, end_column=MOT_START + ncols_mot - 1)
        tm.value = "🏆 TOP Motoristas Ofensores"
        for ci, cn in enumerate(MOT_HDR, MOT_START):
            c = ws.cell(mot_row_start + 1, ci, cn)
            c.fill = hf_mot; c.font = _hfnt(); c.alignment = _CTR; c.border = _BRD
            ws.column_dimensions[get_column_letter(ci)].width = 22

        total_dia_ref = int(r_sta["dia_total"].sum()) if not r_sta.empty else 1
        for ri, (_, row) in enumerate(top5.iterrows(), mot_row_start + 2):
            total_mot = int(row.get("total", 0) or 0)
            pct = total_mot / max(total_dia_ref, 1)
            vals = [
                str(row.get("supervisor", "") or ""),
                row.get("motorista", ""),
                str(row.get("id_motorista", "") or ""),
                str(row.get("ds", "") or ""),
                total_mot,
                pct,
            ]
            for ci, val in enumerate(vals, MOT_START):
                c = ws.cell(ri, ci, val)
                c.font = _bfnt(); c.border = _BRD; c.alignment = _CTR
                if ci in (MOT_START, MOT_START+1): c.alignment = Alignment(horizontal="left", vertical="center")
                if ci == MOT_START + 4:
                    c.number_format = "#,##0"
                    c.font = Font(name="Calibri", size=11, bold=True, color=C_RED)
                if ci == MOT_START + 5:
                    c.number_format = "0.00%"
                    c.fill = PatternFill("solid", fgColor="FF0000") if pct > 0.005 else PatternFill("solid", fgColor="70AD47")
                    c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

    ws.freeze_panes = "A4"

    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Reclamacoes_{data_str}.xlsx"},
    )


# ══════════════════════════════════════════════════════════════
#  TRIAGEM EXCEL — Layout igual ao modelo de referência
#  Dashboard tab: tabela DS (esq) + Top5 (dir) + Total Erro box
#  Abas separadas: Por DS detalhado, Por Supervisor
# ══════════════════════════════════════════════════════════════
@router.get("/triagem/{upload_id}")
@limiter.limit("20/minute")
def excel_triagem(request: Request, upload_id: int, user: dict = Depends(get_current_user)):
    sb = get_supabase()
    upload = sb.table("triagem_uploads").select("*").eq("id", upload_id).execute()
    if not upload.data: raise HTTPException(404, "Não encontrado")
    u = upload.data[0]

    por_ds  = pd.DataFrame(sb.table("triagem_por_ds").select("*").eq("upload_id", upload_id).execute().data or [])
    top5    = pd.DataFrame(sb.table("triagem_top5").select("*").eq("upload_id", upload_id).execute().data or [])
    por_sup = pd.DataFrame(sb.table("triagem_por_supervisor").select("*").eq("upload_id", upload_id).execute().data or [])

    wb = Workbook()

    # ── Cores ──────────────────────────────────────────────────
    C_TITLE  = "1F3864"
    C_HDR_DS = "1F4E79"
    C_HDR_T5 = "C55A11"
    C_HDR_SU = "375623"
    C_OK     = "70AD47"
    C_NOK    = "FF0000"
    C_ALT    = "D9E1F2"
    C_TOTAL  = "BDD7EE"

    def _hfnt(color="FFFFFF"):
        return Font(name="Calibri", bold=True, color=color, size=11)

    def _bfnt(bold=False, color="000000"):
        return Font(name="Calibri", size=10, bold=bold, color=color)

    # ── Aba Dashboard (igual ao modelo) ───────────────────────
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    # Título
    ncols_title = 10
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_title)
    t = ws.cell(1, 1, f"ERRO DE TRIAGEM (DC > DS)  —  {u['data_ref']}")
    t.font = Font(bold=True, color="FFFFFF", name="Calibri", size=16)
    t.fill = PatternFill("solid", fgColor=C_TITLE)
    t.alignment = _CTR
    ws.row_dimensions[1].height = 36

    # ── Tabela DS (colunas 1–6, linha 3+) ─────────────────────
    DS_COLS = ["DS", "Total Expedido", "Triagem OK", "Triagem NOK", "Fora Abrang.", "Taxa (%)"]
    for ci, cn in enumerate(DS_COLS, 1):
        c = ws.cell(3, ci, cn)
        c.fill = PatternFill("solid", fgColor=C_HDR_DS)
        c.font = _hfnt(); c.alignment = _CTR; c.border = _BRD
    ws.row_dimensions[3].height = 24

    if not por_ds.empty:
        df_ds = por_ds.copy()
        df_ds = df_ds.sort_values("taxa" if "taxa" in df_ds.columns else df_ds.columns[0])
        for ri, (_, row) in enumerate(df_ds.iterrows(), 4):
            taxa_val = float(row.get("taxa", 0) or 0)
            vals = [
                row.get("ds", ""),
                int(row.get("total", 0) or 0),
                int(row.get("ok", 0) or 0),
                int(row.get("nok", 0) or 0),
                int(row.get("fora", 0) or 0),
                taxa_val / 100 if taxa_val > 1 else taxa_val,
            ]
            alt = ri % 2 == 0
            for ci, val in enumerate(vals, 1):
                c = ws.cell(ri, ci, val)
                c.font = _bfnt(); c.border = _BRD; c.alignment = _CTR
                if alt: c.fill = PatternFill("solid", fgColor=C_ALT)
                if ci == 1: c.alignment = Alignment(horizontal="left", vertical="center")
                if ci == 6:
                    c.number_format = "0.0%"
                    c.fill = PatternFill("solid", fgColor=C_NOK if taxa_val < 90 else C_OK)
                    c.font = _bfnt(bold=True, color="FFFFFF")
                elif ci in (2, 3, 4, 5):
                    c.number_format = "#,##0"

        # Linha de totais
        last_row = 4 + len(df_ds)
        totals = [
            "Total Geral",
            int(df_ds["total"].sum()) if "total" in df_ds.columns else 0,
            int(df_ds["ok"].sum()) if "ok" in df_ds.columns else 0,
            int(df_ds["nok"].sum()) if "nok" in df_ds.columns else 0,
            int(df_ds.get("fora", pd.Series([0])).sum()),
            u["taxa"] / 100 if float(u["taxa"]) > 1 else float(u["taxa"]),
        ]
        for ci, val in enumerate(totals, 1):
            c = ws.cell(last_row, ci, val)
            c.fill = PatternFill("solid", fgColor=C_TOTAL)
            c.font = _hfnt(color="1F3864"); c.border = _BRD; c.alignment = _CTR
            if ci == 1: c.alignment = Alignment(horizontal="left", vertical="center")
            if ci == 6: c.number_format = "0.0%"
            elif ci in (2, 3, 4, 5): c.number_format = "#,##0"

    # ── Tabela Top 5 (colunas 8–9, linha 3+) ──────────────────
    T5_START = 8
    ws.merge_cells(start_row=2, start_column=T5_START, end_row=2, end_column=T5_START + 1)
    th = ws.cell(2, T5_START, "🏆 TOP 5 DS com mais Erros")
    th.fill = PatternFill("solid", fgColor=C_HDR_T5)
    th.font = _hfnt(); th.alignment = _CTR

    for ci, cn in enumerate(["DS", "Total Erros"], T5_START):
        c = ws.cell(3, ci, cn)
        c.fill = PatternFill("solid", fgColor=C_HDR_T5)
        c.font = _hfnt(); c.alignment = _CTR; c.border = _BRD

    if not top5.empty:
        for ri, (_, row) in enumerate(top5.iterrows(), 4):
            for ci, val in zip([T5_START, T5_START + 1], [row.get("ds", ""), int(row.get("total_erros", 0) or 0)]):
                c = ws.cell(ri, ci, val)
                c.font = _bfnt(bold=(ci == T5_START + 1), color="C00000" if ci == T5_START + 1 else "000000")
                c.border = _BRD; c.alignment = _CTR
                if ci == T5_START: c.alignment = Alignment(horizontal="left", vertical="center")
                if ci == T5_START + 1: c.number_format = "#,##0"

    # ── Box TOTAL ERRO (coluna 8, abaixo do top5) ─────────────
    box_row = 4 + (len(top5) if not top5.empty else 0) + 2
    ws.merge_cells(start_row=box_row, start_column=T5_START, end_row=box_row, end_column=T5_START + 1)
    lbl = ws.cell(box_row, T5_START, "TOTAL ERRO")
    lbl.fill = PatternFill("solid", fgColor=C_HDR_T5); lbl.font = _hfnt(); lbl.alignment = _CTR

    ws.merge_cells(start_row=box_row + 1, start_column=T5_START, end_row=box_row + 1, end_column=T5_START + 1)
    val_cell = ws.cell(box_row + 1, T5_START, int(u["qtd_erro"] or 0))
    val_cell.fill = PatternFill("solid", fgColor=C_NOK)
    val_cell.font = Font(name="Calibri", bold=True, size=20, color="FFFFFF")
    val_cell.alignment = _CTR
    val_cell.number_format = "#,##0"
    ws.row_dimensions[box_row + 1].height = 36

    # ── Larguras ───────────────────────────────────────────────
    for col, w in [(1,22),(2,18),(3,14),(4,16),(5,16),(6,10),(7,4),(8,22),(9,14)]:
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"

    # ── Aba Por DS detalhada ────────────────────────────────────
    if not por_ds.empty:
        ws2 = wb.create_sheet("Por DS")
        df_ds2 = por_ds[["ds", "total", "ok", "nok", "fora", "taxa"]].copy()
        df_ds2.columns = ["DS", "Total Expedido", "Triagem OK", "Triagem NOK", "Fora Abrangência", "Taxa (%)"]
        df_ds2 = df_ds2.sort_values("Taxa (%)")
        _titulo_aba(ws2, f"Resultado por DS — {u['data_ref']}", len(df_ds2.columns))
        _write_header(ws2, df_ds2.columns.tolist(), 2)
        # Escreve com formatação manual para corrigir notação científica
        for ri, (_, row) in enumerate(df_ds2.iterrows(), 3):
            alt = ri % 2 == 0
            for ci, val in enumerate(row, 1):
                try:
                    if pd.isnull(val): val = None
                except: pass
                c = ws2.cell(ri, ci, val)
                c.font = _bfnt(); c.alignment = _CTR; c.border = _BRD
                if alt: c.fill = PatternFill("solid", fgColor=C_ALT)
                if ci == 6: c.number_format = "0.0%"
                elif ci in (2, 3, 4, 5): c.number_format = "#,##0"
        _auto_width(ws2); ws2.freeze_panes = "A3"

    # ── Aba Por Supervisor ─────────────────────────────────────
    if not por_sup.empty:
        ws4 = wb.create_sheet("Por Supervisor")
        df_s = por_sup[["supervisor", "total", "ok", "nok", "fora", "taxa"]].copy()
        df_s.columns = ["Supervisor", "Total", "OK", "NOK", "Fora", "Taxa (%)"]
        df_s = df_s.sort_values("Taxa (%)")
        _titulo_aba(ws4, f"Resultado por Supervisor — {u['data_ref']}", len(df_s.columns))
        _write_header(ws4, df_s.columns.tolist(), 2)
        for ri, (_, row) in enumerate(df_s.iterrows(), 3):
            alt = ri % 2 == 0
            for ci, val in enumerate(row, 1):
                try:
                    if pd.isnull(val): val = None
                except: pass
                c = ws4.cell(ri, ci, val)
                c.font = _bfnt(); c.alignment = _CTR; c.border = _BRD
                if alt: c.fill = PatternFill("solid", fgColor=C_ALT)
                if ci == 6: c.number_format = "0.0%"
                elif ci in (2, 3, 4, 5): c.number_format = "#,##0"
        _auto_width(ws4); ws4.freeze_panes = "A3"

    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Triagem_{u['data_ref']}.xlsx"},
    )


# ══════════════════════════════════════════════════════════════
#  HISTÓRICO EXCEL (período consolidado)
# ══════════════════════════════════════════════════════════════
@router.get("/historico")
@limiter.limit("20/minute")
def excel_historico(
    request: Request,
    data_ini: str = Query(...),
    data_fim: str = Query(...),
    user: dict = Depends(get_current_user),
):
    sb = get_supabase()
    q = (sb.table("expedicao_diaria").select("*")
         .gte("data_ref", data_ini).lte("data_ref", data_fim))
    if user["bases"]: q = q.in_("scan_station", user["bases"])
    data = q.execute().data or []
    if not data: raise HTTPException(404, "Sem dados no período")

    df = pd.DataFrame(data)

    # Cidades do período
    qc = (sb.table("expedicao_cidades").select("*")
          .gte("data_ref", data_ini).lte("data_ref", data_fim))
    if user["bases"]: qc = qc.in_("scan_station", user["bases"])
    cid = pd.DataFrame(qc.execute().data or [])

    # Agrega por DS
    agg = (df.groupby(["scan_station", "region"], as_index=False)
           .agg(recebido=("recebido", "sum"), expedido=("expedido", "sum"),
                entregas=("entregas", "sum"), meta=("meta", "mean")))
    agg["taxa_exp"] = np.where(agg["recebido"] > 0, agg["expedido"] / agg["recebido"], 0)
    agg["taxa_ent"] = np.where(agg["recebido"] > 0, (agg["entregas"] / agg["recebido"]).clip(upper=1.0), 0)
    agg = agg.sort_values("recebido", ascending=False)

    # Agrega cidades
    cid_agg = pd.DataFrame()
    if not cid.empty:
        cid_agg = (cid.groupby(["scan_station", "destination_city"], as_index=False)
                   .agg(recebido=("recebido", "sum"), expedido=("expedido", "sum"),
                        entregas=("entregas", "sum")))
        cid_agg["taxa_exp"] = np.where(cid_agg["recebido"] > 0, cid_agg["expedido"] / cid_agg["recebido"], 0)
        cid_agg["taxa_ent"] = np.where(cid_agg["recebido"] > 0, (cid_agg["entregas"] / cid_agg["recebido"]).clip(upper=1.0), 0)

    wb = Workbook()
    data_str = f"{data_ini} a {data_fim}"

    # Consolidado
    ws = wb.active; ws.title = "Consolidado_Geral"
    _titulo_aba(ws, f"Consolidado Geral — {data_str}", 7)
    _write_grouped(ws, agg, cid_agg, start_row=3)

    # Por região
    for regiao, label in [("capital", "Capital"), ("metropolitan", "Metropolitan"), ("countryside", "Countryside")]:
        df_r = agg[agg["region"].str.lower() == regiao]
        if df_r.empty: continue
        cid_r = cid_agg[cid_agg["scan_station"].isin(df_r["scan_station"])] if not cid_agg.empty else pd.DataFrame()
        ws_r = wb.create_sheet(label)
        _titulo_aba(ws_r, label, 7)
        _write_grouped(ws_r, df_r, cid_r, start_row=3)

    # Resumo por dia
    ws_d = wb.create_sheet("Por Dia")
    dia_agg = (df.groupby("data_ref", as_index=False)
               .agg(recebido=("recebido", "sum"), expedido=("expedido", "sum"), entregas=("entregas", "sum")))
    dia_agg["taxa_exp"] = np.where(dia_agg["recebido"] > 0, dia_agg["expedido"] / dia_agg["recebido"], 0)
    dia_agg = dia_agg.sort_values("data_ref")
    dia_agg.columns = ["Data", "Recebido", "Expedido", "Entregas", "Taxa Exp."]
    _titulo_aba(ws_d, f"Resumo por Dia — {data_str}", len(dia_agg.columns))
    _write_header(ws_d, dia_agg.columns.tolist(), 2)
    _write_data(ws_d, dia_agg, 3, pct_cols=["Taxa Exp."])
    _auto_width(ws_d); ws_d.freeze_panes = "A3"

    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Historico_{data_ini}_a_{data_fim}.xlsx"},
    )


# ══════════════════════════════════════════════════════════════
#  NOT ARRIVED EXCEL (有发未到) — pivot Supervisor×DS×Data
# ══════════════════════════════════════════════════════════════
_SUP_F  = PatternFill("solid", fgColor="1F3864")
_SUP_FN = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_DS_F   = PatternFill("solid", fgColor="DEEAF1")
_DS_FN  = Font(name="Arial", size=10)
_GRD_F  = PatternFill("solid", fgColor="FFC7CE")
_TOT_F  = PatternFill("solid", fgColor="F2F2F2")


def _na_cell_fill(val, max_val):
    """Retorna PatternFill de heatmap para célula com valor > 0."""
    if not val or max_val == 0:
        return None
    r = val / max_val
    if r >= 0.75: return PatternFill("solid", fgColor="FF0000")
    if r >= 0.50: return PatternFill("solid", fgColor="FF9999")
    if r >= 0.25: return PatternFill("solid", fgColor="FFD966")
    return PatternFill("solid", fgColor="FFF2CC")


@router.get("/na/{upload_id}")
@limiter.limit("10/minute")
def excel_na(upload_id: int, request: Request, user: dict = Depends(get_current_user)):
    sb = get_supabase()

    # Busca dados
    up = sb.table("na_uploads").select("*").eq("id", upload_id).execute().data
    if not up:
        raise HTTPException(404, "Upload não encontrado.")
    up = up[0]
    data_ref = up.get("data_ref", "")
    threshold = up.get("threshold_col", ">10D")

    tend = sb.table("na_tendencia").select("supervisor,ds,data,total").eq("upload_id", upload_id).order("data").execute().data or []
    por_sup = sb.table("na_por_supervisor").select("*").eq("upload_id", upload_id).order("total", desc=True).execute().data or []
    por_ds  = sb.table("na_por_ds").select("*").eq("upload_id", upload_id).order("supervisor,total", desc=True).execute().data or []
    por_proc= sb.table("na_por_processo").select("*").eq("upload_id", upload_id).order("total", desc=True).execute().data or []

    # Monta pivot: supervisor → ds → date → count
    pivot = {}
    date_set = set()
    for r in tend:
        s, d, dt, v = r["supervisor"], r["ds"], r["data"], r["total"]
        if s not in pivot: pivot[s] = {}
        if d not in pivot[s]: pivot[s][d] = {}
        pivot[s][d][dt] = pivot[s][d].get(dt, 0) + v
        date_set.add(dt)

    dates = sorted(date_set)
    sup_order = [r["supervisor"] for r in por_sup]
    ds_grd    = {(r["supervisor"], r["ds"]): r.get("grd10d", 0) for r in por_ds}
    sup_info  = {r["supervisor"]: r for r in por_sup}

    max_val = max((r["total"] for r in tend), default=1)

    wb = Workbook()

    # ── Aba "Tendência" (pivot Sheet1) ─────────────────────────
    ws = wb.active
    ws.title = "Tendência"
    n_date_cols = len(dates)
    total_cols  = 3 + n_date_cols + 1  # Supervisor | DS | >10D | dates... | Total

    _titulo_aba(ws, f"Not Arrived — {data_ref}", total_cols)

    # Cabeçalho
    h_row = 2
    headers = ["Supervisor", "DS", threshold] + [dt[5:] for dt in dates] + ["Total"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=h_row, column=ci, value=h)
        c.font = _HFNT; c.fill = _HF; c.alignment = _CTR; c.border = _BRD

    cur = 3
    for sup in sup_order:
        sup_ds_map = pivot.get(sup, {})
        sup_inf    = sup_info.get(sup, {"total": 0, "grd10d": 0})
        ds_list    = sorted(sup_ds_map.keys(), key=lambda d: sup_inf.get("total", 0), reverse=True)

        # Linha supervisor
        ws.cell(row=cur, column=1, value=sup).font = _SUP_FN
        ws.cell(row=cur, column=1).fill = _SUP_F
        ws.cell(row=cur, column=2, value="").fill = _SUP_F
        ws.cell(row=cur, column=3, value=sup_inf.get("grd10d", 0) or None).fill = _SUP_F
        ws.cell(row=cur, column=3).font = Font(name="Arial", bold=True, color="FFC7CE", size=10)
        for ci, dt in enumerate(dates, 4):
            day_tot = sum(sup_ds_map.get(ds, {}).get(dt, 0) for ds in sup_ds_map)
            c = ws.cell(row=cur, column=ci, value=day_tot or None)
            c.fill = _SUP_F; c.font = Font(name="Arial", color="FFFFFF", size=9)
            c.alignment = _CTR; c.border = _BRD
        ws.cell(row=cur, column=3 + n_date_cols + 1, value=sup_inf.get("total", 0)).fill = _SUP_F
        ws.cell(row=cur, column=3 + n_date_cols + 1).font = _SUP_FN
        for ci in [1, 2, 3, 3 + n_date_cols + 1]:
            ws.cell(row=cur, column=ci).alignment = _CTR
            ws.cell(row=cur, column=ci).border = _BRD
        cur += 1

        # Linhas DS
        for ds in ds_list:
            ds_date = sup_ds_map[ds]
            grd     = ds_grd.get((sup, ds), 0)
            ds_tot  = sum(ds_date.values())

            ws.cell(row=cur, column=1, value="").fill = _DS_F
            c2 = ws.cell(row=cur, column=2, value=ds)
            c2.font = _DS_FN; c2.fill = _DS_F; c2.alignment = _LFT; c2.border = _BRD

            c3 = ws.cell(row=cur, column=3, value=grd or None)
            c3.fill = _GRD_F if grd else _DS_F; c3.alignment = _CTR; c3.border = _BRD

            for ci, dt in enumerate(dates, 4):
                val = ds_date.get(dt, 0)
                c = ws.cell(row=cur, column=ci, value=val or None)
                fill = _na_cell_fill(val, max_val)
                if fill: c.fill = fill
                c.alignment = _CTR; c.border = _BRD; c.font = _BFNT

            ct = ws.cell(row=cur, column=3 + n_date_cols + 1, value=ds_tot or None)
            ct.fill = _TOT_F; ct.font = Font(name="Arial", bold=True, size=10)
            ct.alignment = _CTR; ct.border = _BRD
            cur += 1

    # Linha total geral
    cur += 1
    ws.cell(row=cur, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=cur, column=3, value=sum(r.get("grd10d", 0) for r in por_sup) or None).font = Font(name="Arial", bold=True, color="FF0000")
    for ci, dt in enumerate(dates, 4):
        tot_d = sum(pivot.get(s, {}).get(ds, {}).get(dt, 0) for s in pivot for ds in pivot[s])
        ws.cell(row=cur, column=ci, value=tot_d or None)
    ws.cell(row=cur, column=3 + n_date_cols + 1, value=sum(r.get("total", 0) for r in por_sup))
    for ci in range(1, total_cols + 1):
        c = ws.cell(row=cur, column=ci)
        c.fill = _TOT_F; c.border = _BRD; c.alignment = _CTR
        if not c.font.bold:
            c.font = Font(name="Arial", bold=True, size=10)

    ws.freeze_panes = "D3"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 26
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    for ci in range(4, 4 + n_date_cols):
        ws.column_dimensions[get_column_letter(ci)].width = 7
    ws.column_dimensions[get_column_letter(3 + n_date_cols + 1)].width = 8

    # ── Aba "Supervisores" ──────────────────────────────────────
    ws2 = wb.create_sheet("Supervisores")
    _titulo_aba(ws2, f"Por Supervisor — {data_ref}", 4)
    _write_header(ws2, ["Supervisor", threshold, "Total", "% Backlog"], 2)
    df_sup = pd.DataFrame(por_sup)[["supervisor", "grd10d", "total"]] if por_sup else pd.DataFrame(columns=["supervisor","grd10d","total"])
    grand  = df_sup["total"].sum() or 1
    df_sup["pct"] = df_sup["total"] / grand
    df_sup.columns = ["Supervisor", threshold, "Total", "% do Total"]
    _write_data(ws2, df_sup, 3, pct_cols=["% do Total"])
    _auto_width(ws2); ws2.freeze_panes = "A3"

    # ── Aba "Por DS" ────────────────────────────────────────────
    ws3 = wb.create_sheet("Por DS")
    _titulo_aba(ws3, f"Por DS — {data_ref}", 4)
    _write_header(ws3, ["Supervisor", "DS", threshold, "Total"], 2)
    df_ds = pd.DataFrame(por_ds)[["supervisor","ds","grd10d","total"]] if por_ds else pd.DataFrame(columns=["supervisor","ds","grd10d","total"])
    df_ds.columns = ["Supervisor","DS", threshold, "Total"]
    _write_data(ws3, df_ds, 3)
    _auto_width(ws3); ws3.freeze_panes = "A3"

    # ── Aba "Por Processo" ──────────────────────────────────────
    ws4 = wb.create_sheet("Por Processo")
    _titulo_aba(ws4, f"Por Processo — {data_ref}", 2)
    _write_header(ws4, ["Processo", "Total"], 2)
    df_proc = pd.DataFrame(por_proc) if por_proc else pd.DataFrame(columns=["processo","total"])
    df_proc.columns = ["Processo","Total"]
    _write_data(ws4, df_proc, 3)
    _auto_width(ws4); ws4.freeze_panes = "A3"

    fname = f"NotArrived_{data_ref}.xlsx"
    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )

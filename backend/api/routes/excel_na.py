"""
api/routes/excel_na.py — Excel Not Arrived (有发未到)
"""
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from api.deps import get_db, get_current_user
from api.limiter import limiter
from api.routes.excel_base import (
    _HF, _HFNT, _BFNT, _CTR, _LFT, _BRD,
    _titulo_aba, _write_header, _write_data, _auto_width, _to_stream,
)
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

router = APIRouter()

_SUP_F  = PatternFill("solid", fgColor="1F3864")
_SUP_FN = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_DS_F   = PatternFill("solid", fgColor="DEEAF1")
_DS_FN  = Font(name="Arial", size=10)
_GRD_F  = PatternFill("solid", fgColor="FFC7CE")
_TOT_F  = PatternFill("solid", fgColor="F2F2F2")


def _na_cell_fill(val, max_val):
    if not val or max_val == 0:
        return None
    r = val / max_val
    if r >= 0.75: return PatternFill("solid", fgColor="FF0000")
    if r >= 0.50: return PatternFill("solid", fgColor="FF9999")
    if r >= 0.25: return PatternFill("solid", fgColor="FFD966")
    return PatternFill("solid", fgColor="FFF2CC")


@router.get("/na/{upload_id}")
@limiter.limit("10/minute")
def excel_na(upload_id: int, request: Request, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    up_row = db.execute(
        text("SELECT * FROM na_uploads WHERE id = :id"), {"id": upload_id}
    ).mappings().first()
    if not up_row:
        raise HTTPException(404, "Upload não encontrado.")
    up = dict(up_row)
    data_ref  = up.get("data_ref", "")
    threshold = up.get("threshold_col", ">10D")

    tend_rows = db.execute(
        text("SELECT supervisor, ds, data, total FROM na_tendencia WHERE upload_id = :uid ORDER BY data"),
        {"uid": upload_id}
    ).mappings().all()
    tend = [dict(r) for r in tend_rows]

    sup_rows = db.execute(
        text("SELECT supervisor, total, grd10d FROM na_por_supervisor WHERE upload_id = :uid ORDER BY total DESC"),
        {"uid": upload_id}
    ).mappings().all()
    por_sup = [dict(r) for r in sup_rows]

    ds_rows = db.execute(
        text("SELECT supervisor, ds, total, grd10d FROM na_por_ds WHERE upload_id = :uid ORDER BY supervisor, total DESC"),
        {"uid": upload_id}
    ).mappings().all()
    por_ds = [dict(r) for r in ds_rows]

    proc_rows = db.execute(
        text("SELECT processo, total FROM na_por_processo WHERE upload_id = :uid ORDER BY total DESC"),
        {"uid": upload_id}
    ).mappings().all()
    por_proc = [dict(r) for r in proc_rows]

    # Pivot: supervisor → ds → date → count
    pivot = {}
    date_set = set()
    for r in tend:
        s, d, dt, v = r["supervisor"], r["ds"], r["data"], r["total"]
        if s not in pivot: pivot[s] = {}
        if d not in pivot[s]: pivot[s][d] = {}
        pivot[s][d][dt] = pivot[s][d].get(dt, 0) + v
        date_set.add(dt)

    dates     = sorted(date_set)
    sup_order = [r["supervisor"] for r in por_sup]
    ds_grd    = {(r["supervisor"], r["ds"]): r.get("grd10d", 0) for r in por_ds}
    sup_info  = {r["supervisor"]: r for r in por_sup}
    max_val   = max((r["total"] for r in tend), default=1)

    wb = Workbook()

    # ── Aba "Tendência" ───────────────────────────────────────
    ws = wb.active
    ws.title = "Tendência"
    n_date_cols = len(dates)
    total_cols  = 3 + n_date_cols + 1

    _titulo_aba(ws, f"Not Arrived — {data_ref}", total_cols)

    h_row   = 2
    headers = ["Supervisor", "DS", threshold] + [dt[5:] for dt in dates] + ["Total"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=h_row, column=ci, value=h)
        c.font = _HFNT; c.fill = _HF; c.alignment = _CTR; c.border = _BRD

    cur = 3
    for sup in sup_order:
        sup_ds_map = pivot.get(sup, {})
        sup_inf    = sup_info.get(sup, {"total": 0, "grd10d": 0})
        ds_list    = sorted(sup_ds_map.keys(), key=lambda d: sup_inf.get("total", 0), reverse=True)

        ws.cell(row=cur, column=1, value=sup).font = _SUP_FN
        ws.cell(row=cur, column=1).fill = _SUP_F
        ws.cell(row=cur, column=2, value="").fill = _SUP_F
        ws.cell(row=cur, column=3, value=sup_inf.get("grd10d", 0) or None).fill = _SUP_F
        ws.cell(row=cur, column=3).font = Font(name="Arial", bold=True, color="FFC7CE", size=10)
        for ci, dt in enumerate(dates, 4):
            day_tot = sum(sup_ds_map.get(ds, {}).get(dt, 0) for ds in sup_ds_map)
            c = ws.cell(row=cur, column=ci, value=day_tot or None)
            c.fill = _SUP_F; c.font = Font(name="Arial", color="FFFFFF", size=9)
            c.alignment = _CTR; c.border = _BRD
        ws.cell(row=cur, column=3 + n_date_cols + 1, value=sup_inf.get("total", 0)).fill = _SUP_F
        ws.cell(row=cur, column=3 + n_date_cols + 1).font = _SUP_FN
        for ci in [1, 2, 3, 3 + n_date_cols + 1]:
            ws.cell(row=cur, column=ci).alignment = _CTR
            ws.cell(row=cur, column=ci).border = _BRD
        cur += 1

        for ds in ds_list:
            ds_date = sup_ds_map[ds]
            grd     = ds_grd.get((sup, ds), 0)
            ds_tot  = sum(ds_date.values())

            ws.cell(row=cur, column=1, value="").fill = _DS_F
            c2 = ws.cell(row=cur, column=2, value=ds)
            c2.font = _DS_FN; c2.fill = _DS_F; c2.alignment = _LFT; c2.border = _BRD

            c3 = ws.cell(row=cur, column=3, value=grd or None)
            c3.fill = _GRD_F if grd else _DS_F; c3.alignment = _CTR; c3.border = _BRD

            for ci, dt in enumerate(dates, 4):
                val  = ds_date.get(dt, 0)
                c    = ws.cell(row=cur, column=ci, value=val or None)
                fill = _na_cell_fill(val, max_val)
                if fill: c.fill = fill
                c.alignment = _CTR; c.border = _BRD; c.font = _BFNT

            ct = ws.cell(row=cur, column=3 + n_date_cols + 1, value=ds_tot or None)
            ct.fill = _TOT_F; ct.font = Font(name="Arial", bold=True, size=10)
            ct.alignment = _CTR; ct.border = _BRD
            cur += 1

    # Linha total geral
    cur += 1
    ws.cell(row=cur, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=cur, column=3, value=sum(r.get("grd10d", 0) for r in por_sup) or None).font = Font(name="Arial", bold=True, color="FF0000")
    for ci, dt in enumerate(dates, 4):
        tot_d = sum(pivot.get(s, {}).get(ds, {}).get(dt, 0) for s in pivot for ds in pivot[s])
        ws.cell(row=cur, column=ci, value=tot_d or None)
    ws.cell(row=cur, column=3 + n_date_cols + 1, value=sum(r.get("total", 0) for r in por_sup))
    for ci in range(1, total_cols + 1):
        c = ws.cell(row=cur, column=ci)
        c.fill = _TOT_F; c.border = _BRD; c.alignment = _CTR
        if not c.font.bold:
            c.font = Font(name="Arial", bold=True, size=10)

    ws.freeze_panes = "D3"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 26
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    for ci in range(4, 4 + n_date_cols):
        ws.column_dimensions[get_column_letter(ci)].width = 7
    ws.column_dimensions[get_column_letter(3 + n_date_cols + 1)].width = 8

    # ── Aba "Supervisores" ────────────────────────────────────
    ws2 = wb.create_sheet("Supervisores")
    _titulo_aba(ws2, f"Por Supervisor — {data_ref}", 4)
    _write_header(ws2, ["Supervisor", threshold, "Total", "% Backlog"], 2)
    df_sup = pd.DataFrame(por_sup if por_sup else [], columns=["supervisor", "grd10d", "total"])
    if por_sup:
        df_sup = df_sup[["supervisor", "grd10d", "total"]]
    grand = int(df_sup["total"].sum()) or 1
    df_sup = df_sup.copy()
    df_sup["pct"] = df_sup["total"] / grand
    df_sup.columns = ["Supervisor", threshold, "Total", "% do Total"]
    _write_data(ws2, df_sup, 3, pct_cols=["% do Total"])
    _auto_width(ws2); ws2.freeze_panes = "A3"

    # ── Aba "Por DS" ──────────────────────────────────────────
    ws3 = wb.create_sheet("Por DS")
    _titulo_aba(ws3, f"Por DS — {data_ref}", 4)
    _write_header(ws3, ["Supervisor", "DS", threshold, "Total"], 2)
    df_ds = pd.DataFrame(por_ds if por_ds else [], columns=["supervisor", "ds", "grd10d", "total"])
    if por_ds:
        df_ds = df_ds[["supervisor", "ds", "grd10d", "total"]]
    df_ds.columns = ["Supervisor", "DS", threshold, "Total"]
    _write_data(ws3, df_ds, 3)
    _auto_width(ws3); ws3.freeze_panes = "A3"

    # ── Aba "Por Processo" ────────────────────────────────────
    ws4 = wb.create_sheet("Por Processo")
    _titulo_aba(ws4, f"Por Processo — {data_ref}", 2)
    _write_header(ws4, ["Processo", "Total"], 2)
    df_proc = pd.DataFrame(por_proc if por_proc else [], columns=["processo", "total"])
    if por_proc:
        df_proc = df_proc[["processo", "total"]]
    df_proc.columns = ["Processo", "Total"]
    _write_data(ws4, df_proc, 3)
    _auto_width(ws4); ws4.freeze_panes = "A3"

    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=NotArrived_{data_ref}.xlsx"},
    )

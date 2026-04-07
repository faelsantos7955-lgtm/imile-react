"""
api/routes/excel_historico.py — Excel do Histórico (período consolidado)
"""
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from api.deps import get_db, get_current_user
from api.limiter import limiter
from api.routes.excel_base import (
    _titulo_aba, _write_header, _write_data, _write_grouped, _auto_width, _to_stream,
)
import pandas as pd
import numpy as np
from openpyxl import Workbook

router = APIRouter()


@router.get("/historico")
@limiter.limit("20/minute")
def excel_historico(
    request: Request,
    data_ini: str = Query(...),
    data_fim: str = Query(...),
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user["bases"]:
        rows = db.execute(
            text("SELECT * FROM expedicao_diaria WHERE data_ref >= :ini AND data_ref <= :fim AND scan_station = ANY(:bases)"),
            {"ini": data_ini, "fim": data_fim, "bases": user["bases"]}
        ).mappings().all()
        rows_c = db.execute(
            text("SELECT * FROM expedicao_cidades WHERE data_ref >= :ini AND data_ref <= :fim AND scan_station = ANY(:bases)"),
            {"ini": data_ini, "fim": data_fim, "bases": user["bases"]}
        ).mappings().all()
    else:
        rows = db.execute(
            text("SELECT * FROM expedicao_diaria WHERE data_ref >= :ini AND data_ref <= :fim"),
            {"ini": data_ini, "fim": data_fim}
        ).mappings().all()
        rows_c = db.execute(
            text("SELECT * FROM expedicao_cidades WHERE data_ref >= :ini AND data_ref <= :fim"),
            {"ini": data_ini, "fim": data_fim}
        ).mappings().all()

    if not rows:
        raise HTTPException(404, "Sem dados no período")

    df = pd.DataFrame([dict(r) for r in rows])
    cid = pd.DataFrame([dict(r) for r in rows_c])

    agg = (df.groupby(["scan_station", "region"], as_index=False)
           .agg(recebido=("recebido", "sum"), expedido=("expedido", "sum"),
                entregas=("entregas", "sum"), meta=("meta", "mean")))
    agg["taxa_exp"] = np.where(agg["recebido"] > 0, agg["expedido"] / agg["recebido"], 0)
    agg["taxa_ent"] = np.where(agg["recebido"] > 0, (agg["entregas"] / agg["recebido"]).clip(upper=1.0), 0)
    agg = agg.sort_values("recebido", ascending=False)

    cid_agg = pd.DataFrame()
    if not cid.empty:
        cid_agg = (cid.groupby(["scan_station", "destination_city"], as_index=False)
                   .agg(recebido=("recebido", "sum"), expedido=("expedido", "sum"),
                        entregas=("entregas", "sum")))
        cid_agg["taxa_exp"] = np.where(cid_agg["recebido"] > 0, cid_agg["expedido"] / cid_agg["recebido"], 0)
        cid_agg["taxa_ent"] = np.where(cid_agg["recebido"] > 0, (cid_agg["entregas"] / cid_agg["recebido"]).clip(upper=1.0), 0)

    wb = Workbook()
    data_str = f"{data_ini} a {data_fim}"

    ws = wb.active; ws.title = "Consolidado_Geral"
    _titulo_aba(ws, f"Consolidado Geral — {data_str}", 7)
    _write_grouped(ws, agg, cid_agg, start_row=3)

    for regiao, label in [("capital", "Capital"), ("metropolitan", "Metropolitan"), ("countryside", "Countryside")]:
        df_r = agg[agg["region"].str.lower() == regiao]
        if df_r.empty: continue
        cid_r = cid_agg[cid_agg["scan_station"].isin(df_r["scan_station"])] if not cid_agg.empty else pd.DataFrame()
        ws_r = wb.create_sheet(label)
        _titulo_aba(ws_r, label, 7)
        _write_grouped(ws_r, df_r, cid_r, start_row=3)

    ws_d = wb.create_sheet("Por Dia")
    dia_agg = (df.groupby("data_ref", as_index=False)
               .agg(recebido=("recebido", "sum"), expedido=("expedido", "sum"), entregas=("entregas", "sum")))
    dia_agg["taxa_exp"] = np.where(dia_agg["recebido"] > 0, dia_agg["expedido"] / dia_agg["recebido"], 0)
    dia_agg = dia_agg.sort_values("data_ref")
    dia_agg.columns = ["Data", "Recebido", "Expedido", "Entregas", "Taxa Exp."]
    _titulo_aba(ws_d, f"Resumo por Dia — {data_str}", len(dia_agg.columns))
    _write_header(ws_d, dia_agg.columns.tolist(), 2)
    _write_data(ws_d, dia_agg, 3, pct_cols=["Taxa Exp."])
    _auto_width(ws_d); ws_d.freeze_panes = "A3"

    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Historico_{data_ini}_a_{data_fim}.xlsx"},
    )

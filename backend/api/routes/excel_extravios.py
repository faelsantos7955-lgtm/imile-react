"""
api/routes/excel_extravios.py — Excel Controle de Extravios
"""
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from api.deps import get_db, get_current_user
from api.limiter import limiter
from api.routes.excel_base import (
    _HF, _HFNT, _BFNT, _CTR, _BRD, _AF,
    _titulo_aba, _write_header, _write_data, _auto_width, _to_stream,
)
import pandas as pd
from openpyxl import Workbook

router = APIRouter()


@router.get("/extravios/{upload_id}")
@limiter.limit("10/minute")
def excel_extravios(
    upload_id: int, request: Request,
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    up_row = db.execute(
        text("SELECT * FROM extravios_uploads WHERE id = :id"), {"id": upload_id}
    ).mappings().first()
    if not up_row:
        raise HTTPException(404, "Upload não encontrado.")
    up = dict(up_row)
    data_ref = up.get("data_ref", "")

    ds_rows = db.execute(
        text("SELECT ds, supervisor, regional, total, valor_total, total_lost, total_damaged FROM extravios_por_ds WHERE upload_id = :uid ORDER BY total DESC"),
        {"uid": upload_id}
    ).mappings().all()
    por_ds = [dict(r) for r in ds_rows]

    mot_rows = db.execute(
        text("SELECT motivo, total, valor_total FROM extravios_por_motivo WHERE upload_id = :uid ORDER BY total DESC"),
        {"uid": upload_id}
    ).mappings().all()
    por_mot = [dict(r) for r in mot_rows]

    sem_rows = db.execute(
        text("SELECT semana, mes, total, valor_total FROM extravios_por_semana WHERE upload_id = :uid ORDER BY semana"),
        {"uid": upload_id}
    ).mappings().all()
    por_sem = [dict(r) for r in sem_rows]

    wb = Workbook()

    # ── Aba "Por DS" ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Por DS"
    _titulo_aba(ws, f"Extravios — Por DS — {data_ref}", 7)
    _write_header(ws, ["#", "DS", "Supervisor", "Regional", "Total", "Goods Lost", "Avaria", "Valor Declarado"], 2)
    rows = []
    for i, d in enumerate(por_ds, 1):
        rows.append({
            "#": i,
            "DS": d["ds"],
            "Supervisor": d["supervisor"] or "—",
            "Regional": d["regional"] or "—",
            "Total": d["total"],
            "Goods Lost": d["total_lost"],
            "Avaria": d["total_damaged"],
            "Valor Declarado": d["valor_total"],
        })
    df_ds = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["#", "DS", "Supervisor", "Regional", "Total", "Goods Lost", "Avaria", "Valor Declarado"])
    _write_data(ws, df_ds, 3)
    _auto_width(ws)
    ws.freeze_panes = "A3"

    # ── Aba "Por Motivo" ──────────────────────────────────────
    ws2 = wb.create_sheet("Por Motivo")
    _titulo_aba(ws2, f"Extravios — Por Motivo — {data_ref}", 3)
    _write_header(ws2, ["Motivo", "Total", "Valor Declarado"], 2)
    rows2 = [{"Motivo": d["motivo"], "Total": d["total"], "Valor Declarado": d["valor_total"]} for d in por_mot]
    df_mot = pd.DataFrame(rows2) if rows2 else pd.DataFrame(columns=["Motivo", "Total", "Valor Declarado"])
    _write_data(ws2, df_mot, 3)
    _auto_width(ws2)
    ws2.freeze_panes = "A3"

    # ── Aba "Por Semana" ──────────────────────────────────────
    ws3 = wb.create_sheet("Por Semana")
    _titulo_aba(ws3, f"Extravios — Por Semana — {data_ref}", 4)
    _write_header(ws3, ["Semana", "Mês", "Total", "Valor Declarado"], 2)
    rows3 = [{"Semana": d["semana"], "Mês": d["mes"], "Total": d["total"], "Valor Declarado": d["valor_total"]} for d in por_sem]
    df_sem = pd.DataFrame(rows3) if rows3 else pd.DataFrame(columns=["Semana", "Mês", "Total", "Valor Declarado"])
    _write_data(ws3, df_sem, 3)
    _auto_width(ws3)
    ws3.freeze_panes = "A3"

    fname = f"Extravios_{data_ref}.xlsx"
    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )

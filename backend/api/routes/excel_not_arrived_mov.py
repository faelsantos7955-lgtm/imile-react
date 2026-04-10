"""
api/routes/excel_not_arrived_mov.py — Excel Not Arrived com Movimentação
"""
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from api.deps import get_db, get_current_user
from api.limiter import limiter
from api.routes.excel_base import (
    _HF, _HFNT, _BFNT, _CTR, _BRD, _AF,
    _titulo_aba, _write_header, _write_data, _auto_width, _to_stream,
)
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

router = APIRouter()

_SUP_F   = PatternFill("solid", fgColor="1F3864")
_SUP_FN  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_TOT_F   = PatternFill("solid", fgColor="F2F2F2")
_CELL_FILLS = [
    (0.75, PatternFill("solid", fgColor="FF0000")),
    (0.50, PatternFill("solid", fgColor="FF9999")),
    (0.25, PatternFill("solid", fgColor="FFD966")),
    (0.00, PatternFill("solid", fgColor="FFF2CC")),
]

def _mov_cell_fill(val, max_val):
    if not val or max_val == 0:
        return None
    r = val / max_val
    for thr, fill in _CELL_FILLS:
        if r >= thr:
            return fill
    return None


@router.get("/not-arrived-mov/{upload_id}")
@limiter.limit("10/minute")
def excel_not_arrived_mov(
    upload_id: int, request: Request,
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    up_row = db.execute(
        text("SELECT * FROM not_arrived_uploads WHERE id = :id"), {"id": upload_id}
    ).mappings().first()
    if not up_row:
        raise HTTPException(404, "Upload não encontrado.")
    up = dict(up_row)
    data_ref = up.get("data_ref", "")

    # Tendência para aba 汇总
    tend_rows = db.execute(
        text("SELECT supervisor, data, total FROM not_arrived_tendencia WHERE upload_id = :uid ORDER BY data"),
        {"uid": upload_id}
    ).mappings().all()
    tendencia = [dict(r) for r in tend_rows]

    est_rows = db.execute(
        text("SELECT oc_name, oc_code, tipo, regiao, supervisor, total, entregues FROM not_arrived_por_estacao WHERE upload_id = :uid ORDER BY total DESC"),
        {"uid": upload_id}
    ).mappings().all()
    por_estacao = [dict(r) for r in est_rows]

    sup_rows = db.execute(
        text("SELECT supervisor, total, total_dc, total_ds, entregues FROM not_arrived_por_supervisor WHERE upload_id = :uid ORDER BY total DESC"),
        {"uid": upload_id}
    ).mappings().all()
    por_sup = [dict(r) for r in sup_rows]

    reg_rows = db.execute(
        text("SELECT regiao, tipo, total FROM not_arrived_por_regiao WHERE upload_id = :uid ORDER BY total DESC"),
        {"uid": upload_id}
    ).mappings().all()
    por_reg = [dict(r) for r in reg_rows]

    op_rows = db.execute(
        text("SELECT operacao, total FROM not_arrived_por_operacao WHERE upload_id = :uid ORDER BY total DESC"),
        {"uid": upload_id}
    ).mappings().all()
    por_op = [dict(r) for r in op_rows]

    wb = Workbook()

    # ── Aba "汇总" — Dash: Supervisor × Data ──────────────────
    ws_hui = wb.active
    ws_hui.title = "汇总"

    if tendencia:
        pivot = {}
        date_set = set()
        for r in tendencia:
            s, dt, v = r["supervisor"], r["data"], r["total"]
            if s not in pivot: pivot[s] = {}
            pivot[s][dt] = pivot[s].get(dt, 0) + v
            date_set.add(dt)

        dates     = sorted(date_set)
        sup_order = sorted(pivot.keys(), key=lambda s: sum(pivot[s].values()), reverse=True)
        max_val   = max((r["total"] for r in tendencia), default=1)
        n_cols    = len(dates)

        # Título
        _titulo_aba(ws_hui, f"DS端有发未到问题件后又有操作 — {data_ref}", n_cols + 2)

        # Header
        h_row = 2
        ws_hui.cell(row=h_row, column=1, value="区域/Supervisor").font = _HFNT
        ws_hui.cell(row=h_row, column=1).fill = _HF
        ws_hui.cell(row=h_row, column=1).alignment = _CTR
        ws_hui.cell(row=h_row, column=1).border = _BRD
        for ci, dt in enumerate(dates, 2):
            c = ws_hui.cell(row=h_row, column=ci, value=dt[5:].replace("-", "/"))
            c.font = _HFNT; c.fill = _HF; c.alignment = _CTR; c.border = _BRD
        ct = ws_hui.cell(row=h_row, column=n_cols + 2, value="Total")
        ct.font = _HFNT; ct.fill = _HF; ct.alignment = _CTR; ct.border = _BRD

        cur = 3
        for sup in sup_order:
            sup_data = pivot[sup]
            total_sup = sum(sup_data.values())
            cs = ws_hui.cell(row=cur, column=1, value=sup)
            cs.font = _SUP_FN; cs.fill = _SUP_F; cs.alignment = _CTR; cs.border = _BRD
            for ci, dt in enumerate(dates, 2):
                val = sup_data.get(dt, 0)
                c = ws_hui.cell(row=cur, column=ci, value=val or None)
                fill = _mov_cell_fill(val, max_val)
                if fill: c.fill = fill
                else: c.fill = _SUP_F
                c.font = Font(name="Arial", color="FFFFFF", size=9, bold=bool(fill and val))
                c.alignment = _CTR; c.border = _BRD
            ct2 = ws_hui.cell(row=cur, column=n_cols + 2, value=total_sup)
            ct2.font = _SUP_FN; ct2.fill = _SUP_F; ct2.alignment = _CTR; ct2.border = _BRD
            cur += 1

        # Linha total
        cur += 1
        ws_hui.cell(row=cur, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
        for ci, dt in enumerate(dates, 2):
            tot_d = sum(pivot[s].get(dt, 0) for s in pivot)
            c = ws_hui.cell(row=cur, column=ci, value=tot_d or None)
            c.fill = _TOT_F; c.alignment = _CTR; c.border = _BRD
            c.font = Font(name="Arial", bold=True, size=9)
        grand = sum(sum(pivot[s].values()) for s in pivot)
        ws_hui.cell(row=cur, column=n_cols + 2, value=grand).font = Font(name="Arial", bold=True, size=10)
        for ci in range(1, n_cols + 3):
            ws_hui.cell(row=cur, column=ci).fill = _TOT_F
            ws_hui.cell(row=cur, column=ci).border = _BRD
            ws_hui.cell(row=cur, column=ci).alignment = _CTR

        ws_hui.freeze_panes = "B3"
        ws_hui.column_dimensions["A"].width = 18
        for ci in range(2, n_cols + 2):
            ws_hui.column_dimensions[get_column_letter(ci)].width = 7
        ws_hui.column_dimensions[get_column_letter(n_cols + 2)].width = 8

    # ── Aba "Por Estação" ─────────────────────────────────────
    ws = wb.create_sheet("Por Estação")
    _titulo_aba(ws, f"Not Arrived Mov. — Por Estação — {data_ref}", 8)
    _write_header(ws, ["#", "Estação", "Código", "Tipo", "Região", "Supervisor", "Total", "Entregues"], 2)
    rows = []
    for i, d in enumerate(por_estacao, 1):
        pct_ent = round(d["entregues"] / d["total"] * 100, 1) if d["total"] else 0
        rows.append({
            "#": i,
            "Estação": d["oc_name"] or "—",
            "Código": d["oc_code"] or "—",
            "Tipo": d["tipo"] or "—",
            "Região": d["regiao"] or "—",
            "Supervisor": d["supervisor"] or "—",
            "Total": d["total"],
            "Entregues": d["entregues"],
        })
    df_est = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["#", "Estação", "Código", "Tipo", "Região", "Supervisor", "Total", "Entregues"])
    _write_data(ws, df_est, 3)
    _auto_width(ws)
    ws.freeze_panes = "A3"

    # ── Aba "Por Supervisor" ──────────────────────────────────
    ws2 = wb.create_sheet("Por Supervisor")
    _titulo_aba(ws2, f"Not Arrived Mov. — Por Supervisor — {data_ref}", 5)
    _write_header(ws2, ["Supervisor", "Total", "Total DC", "Total DS", "Entregues"], 2)
    rows2 = []
    for d in por_sup:
        rows2.append({
            "Supervisor": d["supervisor"] or "—",
            "Total": d["total"],
            "Total DC": d["total_dc"],
            "Total DS": d["total_ds"],
            "Entregues": d["entregues"],
        })
    df_sup = pd.DataFrame(rows2) if rows2 else pd.DataFrame(columns=["Supervisor", "Total", "Total DC", "Total DS", "Entregues"])
    _write_data(ws2, df_sup, 3)
    _auto_width(ws2)
    ws2.freeze_panes = "A3"

    # ── Aba "Por Região" ──────────────────────────────────────
    ws3 = wb.create_sheet("Por Região")
    _titulo_aba(ws3, f"Not Arrived Mov. — Por Região — {data_ref}", 3)
    _write_header(ws3, ["Região", "Tipo", "Total"], 2)
    rows3 = [{"Região": d["regiao"], "Tipo": d["tipo"], "Total": d["total"]} for d in por_reg]
    df_reg = pd.DataFrame(rows3) if rows3 else pd.DataFrame(columns=["Região", "Tipo", "Total"])
    _write_data(ws3, df_reg, 3)
    _auto_width(ws3)
    ws3.freeze_panes = "A3"

    # ── Aba "Por Operação" ────────────────────────────────────
    ws4 = wb.create_sheet("Por Operação")
    _titulo_aba(ws4, f"Not Arrived Mov. — Por Operação — {data_ref}", 2)
    _write_header(ws4, ["Operação", "Total"], 2)
    rows4 = [{"Operação": d["operacao"], "Total": d["total"]} for d in por_op]
    df_op = pd.DataFrame(rows4) if rows4 else pd.DataFrame(columns=["Operação", "Total"])
    _write_data(ws4, df_op, 3)
    _auto_width(ws4)
    ws4.freeze_panes = "A3"

    fname = f"NotArrivedMov_{data_ref}.xlsx"
    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )

"""
api/routes/excel_triagem.py — Excel de Triagem (DC > DS)
"""
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from api.deps import get_db, get_current_user
from api.limiter import limiter
from api.routes.excel_base import (
    _CTR, _BRD, _titulo_aba, _write_header, _auto_width, _to_stream,
)
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

router = APIRouter()


@router.get("/triagem/{upload_id}")
@limiter.limit("20/minute")
def excel_triagem(request: Request, upload_id: int, user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    up_row = db.execute(
        text("SELECT * FROM triagem_uploads WHERE id = :id"), {"id": upload_id}
    ).mappings().first()
    if not up_row:
        raise HTTPException(404, "Não encontrado")
    u = dict(up_row)

    por_ds = pd.DataFrame([dict(r) for r in db.execute(
        text("SELECT * FROM triagem_por_ds WHERE upload_id = :uid"), {"uid": upload_id}
    ).mappings().all()])
    top5 = pd.DataFrame([dict(r) for r in db.execute(
        text("SELECT * FROM triagem_top5 WHERE upload_id = :uid"), {"uid": upload_id}
    ).mappings().all()])
    por_sup = pd.DataFrame([dict(r) for r in db.execute(
        text("SELECT * FROM triagem_por_supervisor WHERE upload_id = :uid"), {"uid": upload_id}
    ).mappings().all()])

    wb = Workbook()

    C_TITLE  = "1F3864"
    C_HDR_DS = "1F4E79"
    C_HDR_T5 = "C55A11"
    C_HDR_SU = "375623"
    C_OK     = "70AD47"
    C_NOK    = "FF0000"
    C_ALT    = "D9E1F2"
    C_TOTAL  = "BDD7EE"
    C_REC    = "2E75B6"   # azul — recebidos
    C_ERR    = "C00000"   # vermelho escuro — chegou errado

    tem_arrival = bool(u.get("tem_arrival", False))
    # T5_START: quando tem_arrival=True a tabela DS usa cols 1-8, então Top5 começa na 10
    T5_START = 10 if tem_arrival else 8

    def _hfnt(color="FFFFFF"): return Font(name="Calibri", bold=True, color=color, size=11)
    def _bfnt(bold=False, color="000000"): return Font(name="Calibri", size=10, bold=bold, color=color)

    # ── Aba Dashboard ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    ncols_title = T5_START + 3 if tem_arrival else T5_START + 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols_title)
    t = ws.cell(1, 1, f"ERRO DE TRIAGEM (DC > DS)  —  {u['data_ref']}")
    t.font = Font(bold=True, color="FFFFFF", name="Calibri", size=16)
    t.fill = PatternFill("solid", fgColor=C_TITLE)
    t.alignment = _CTR; ws.row_dimensions[1].height = 36

    # ── Tabela DS ─────────────────────────────────────────────
    DS_COLS = ["DS", "Total Expedido", "Triagem OK", "Triagem NOK", "Fora Abrang.", "Taxa (%)"]
    if tem_arrival:
        DS_COLS += ["Recebidos", "Chegou Errado"]
    for ci, cn in enumerate(DS_COLS, 1):
        c = ws.cell(3, ci, cn)
        fgColor = C_HDR_DS
        if tem_arrival and cn == "Chegou Errado": fgColor = "C00000"
        elif tem_arrival and cn == "Recebidos":   fgColor = "1F4E79"
        c.fill = PatternFill("solid", fgColor=fgColor)
        c.font = _hfnt(); c.alignment = _CTR; c.border = _BRD
    ws.row_dimensions[3].height = 24

    if not por_ds.empty:
        df_ds = por_ds.copy()
        df_ds = df_ds.sort_values("taxa" if "taxa" in df_ds.columns else df_ds.columns[0])
        for ri, (_, row) in enumerate(df_ds.iterrows(), 4):
            taxa_val = float(row.get("taxa", 0) or 0)
            vals = [
                row.get("ds", ""),
                int(row.get("total", 0) or 0),
                int(row.get("ok", 0) or 0),
                int(row.get("nok", 0) or 0),
                int(row.get("fora", 0) or 0),
                taxa_val / 100 if taxa_val > 1 else taxa_val,
            ]
            if tem_arrival:
                vals += [
                    int(row.get("recebidos", 0) or 0),
                    int(row.get("recebidos_nok", 0) or 0),
                ]
            alt = ri % 2 == 0
            for ci, val in enumerate(vals, 1):
                c = ws.cell(ri, ci, val)
                c.font = _bfnt(); c.border = _BRD; c.alignment = _CTR
                if alt: c.fill = PatternFill("solid", fgColor=C_ALT)
                if ci == 1: c.alignment = Alignment(horizontal="left", vertical="center")
                if ci == 6:
                    c.number_format = "0.0%"
                    c.fill = PatternFill("solid", fgColor=C_NOK if taxa_val < 90 else C_OK)
                    c.font = _bfnt(bold=True, color="FFFFFF")
                elif ci in (2, 3, 4, 5): c.number_format = "#,##0"
                elif tem_arrival and ci == 7: c.number_format = "#,##0"
                elif tem_arrival and ci == 8:
                    c.number_format = "#,##0"
                    if int(val) > 0:
                        c.font = _bfnt(bold=True, color="C00000")

        last_row = 4 + len(df_ds)
        totals = [
            "Total Geral",
            int(df_ds["total"].sum()) if "total" in df_ds.columns else 0,
            int(df_ds["ok"].sum()) if "ok" in df_ds.columns else 0,
            int(df_ds["nok"].sum()) if "nok" in df_ds.columns else 0,
            int(df_ds.get("fora", pd.Series([0])).sum()),
            u["taxa"] / 100 if float(u["taxa"]) > 1 else float(u["taxa"]),
        ]
        if tem_arrival:
            totals += [
                int(df_ds["recebidos"].sum()) if "recebidos" in df_ds.columns else 0,
                int(df_ds["recebidos_nok"].sum()) if "recebidos_nok" in df_ds.columns else 0,
            ]
        for ci, val in enumerate(totals, 1):
            c = ws.cell(last_row, ci, val)
            c.fill = PatternFill("solid", fgColor=C_TOTAL)
            c.font = _hfnt(color="1F3864"); c.border = _BRD; c.alignment = _CTR
            if ci == 1: c.alignment = Alignment(horizontal="left", vertical="center")
            if ci == 6: c.number_format = "0.0%"
            elif ci in (2, 3, 4, 5): c.number_format = "#,##0"
            elif tem_arrival and ci == 7: c.number_format = "#,##0"
            elif tem_arrival and ci == 8:
                c.number_format = "#,##0"
                c.fill = PatternFill("solid", fgColor="C00000")
                c.font = _hfnt(color="FFFFFF")

    # ── Tabela Top 5 ─────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=T5_START, end_row=2, end_column=T5_START + 1)
    th = ws.cell(2, T5_START, "🏆 TOP 5 DS com mais Erros")
    th.fill = PatternFill("solid", fgColor=C_HDR_T5)
    th.font = _hfnt(); th.alignment = _CTR

    for ci, cn in enumerate(["DS", "Total Erros"], T5_START):
        c = ws.cell(3, ci, cn)
        c.fill = PatternFill("solid", fgColor=C_HDR_T5)
        c.font = _hfnt(); c.alignment = _CTR; c.border = _BRD

    if not top5.empty:
        for ri, (_, row) in enumerate(top5.iterrows(), 4):
            for ci, val in zip([T5_START, T5_START + 1], [row.get("ds", ""), int(row.get("total_erros", 0) or 0)]):
                c = ws.cell(ri, ci, val)
                c.font = _bfnt(bold=(ci == T5_START + 1), color="C00000" if ci == T5_START + 1 else "000000")
                c.border = _BRD; c.alignment = _CTR
                if ci == T5_START: c.alignment = Alignment(horizontal="left", vertical="center")
                if ci == T5_START + 1: c.number_format = "#,##0"

    # ── Box TOTAL ERRO ────────────────────────────────────────
    box_row = 4 + (len(top5) if not top5.empty else 0) + 2
    ws.merge_cells(start_row=box_row, start_column=T5_START, end_row=box_row, end_column=T5_START + 1)
    lbl = ws.cell(box_row, T5_START, "TOTAL ERRO")
    lbl.fill = PatternFill("solid", fgColor=C_HDR_T5); lbl.font = _hfnt(); lbl.alignment = _CTR

    ws.merge_cells(start_row=box_row + 1, start_column=T5_START, end_row=box_row + 1, end_column=T5_START + 1)
    val_cell = ws.cell(box_row + 1, T5_START, int(u["qtd_erro"] or 0))
    val_cell.fill = PatternFill("solid", fgColor=C_NOK)
    val_cell.font = Font(name="Calibri", bold=True, size=20, color="FFFFFF")
    val_cell.alignment = _CTR; val_cell.number_format = "#,##0"
    ws.row_dimensions[box_row + 1].height = 36

    # Box "Chegou Errado" (apenas quando tem Arrival)
    if tem_arrival and not por_ds.empty:
        total_chegou_errado = int(por_ds["recebidos_nok"].sum()) if "recebidos_nok" in por_ds.columns else 0
        box2_col = T5_START + 3
        ws.merge_cells(start_row=box_row, start_column=box2_col, end_row=box_row, end_column=box2_col + 1)
        lbl2 = ws.cell(box_row, box2_col, "CHEGOU ERRADO NA DS")
        lbl2.fill = PatternFill("solid", fgColor="C00000"); lbl2.font = _hfnt(); lbl2.alignment = _CTR

        ws.merge_cells(start_row=box_row + 1, start_column=box2_col, end_row=box_row + 1, end_column=box2_col + 1)
        val2 = ws.cell(box_row + 1, box2_col, total_chegou_errado)
        val2.fill = PatternFill("solid", fgColor="C00000")
        val2.font = Font(name="Calibri", bold=True, size=20, color="FFFFFF")
        val2.alignment = _CTR; val2.number_format = "#,##0"

    if tem_arrival:
        # DS usa cols 1-8; spacer col 9; Top5 cols 10-11; spacer col 12; box2 cols 13-14
        _col_ws = [(1,22),(2,18),(3,14),(4,16),(5,16),(6,10),(7,16),(8,18),
                   (9,4),(10,22),(11,14),(12,4),(13,22),(14,18)]
    else:
        # DS usa cols 1-6; spacer col 7; Top5 cols 8-9
        _col_ws = [(1,22),(2,18),(3,14),(4,16),(5,16),(6,10),(7,4),(8,22),(9,14)]
    for col, w in _col_ws:
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"

    # ── Aba Por DS detalhada ──────────────────────────────────
    if not por_ds.empty:
        ws2 = wb.create_sheet("Por DS")
        cols_ds2 = ["ds", "total", "ok", "nok", "fora", "taxa"]
        col_names2 = ["DS", "Total Expedido", "Triagem OK", "Triagem NOK", "Fora Abrangência", "Taxa (%)"]
        if tem_arrival:
            if "recebidos" in por_ds.columns:
                cols_ds2.append("recebidos"); col_names2.append("Recebidos")
            if "recebidos_nok" in por_ds.columns:
                cols_ds2.append("recebidos_nok"); col_names2.append("Chegou Errado na DS")
        df_ds2 = por_ds[[c for c in cols_ds2 if c in por_ds.columns]].copy()
        df_ds2.columns = col_names2[:len(df_ds2.columns)]
        df_ds2 = df_ds2.sort_values("Taxa (%)")
        _titulo_aba(ws2, f"Resultado por DS — {u['data_ref']}", len(df_ds2.columns))
        _write_header(ws2, df_ds2.columns.tolist(), 2)
        for ri, (_, row) in enumerate(df_ds2.iterrows(), 3):
            alt = ri % 2 == 0
            for ci, val in enumerate(row, 1):
                try:
                    if pd.isnull(val): val = None
                except: pass
                c = ws2.cell(ri, ci, val)
                c.font = Font(name="Calibri", size=10); c.alignment = _CTR; c.border = _BRD
                if alt: c.fill = PatternFill("solid", fgColor=C_ALT)
                if ci == 6: c.number_format = "0.0%"
                elif ci in (2, 3, 4, 5): c.number_format = "#,##0"
        _auto_width(ws2); ws2.freeze_panes = "A3"

    # ── Aba Por Supervisor ────────────────────────────────────
    if not por_sup.empty:
        ws4 = wb.create_sheet("Por Supervisor")
        df_s = por_sup[["supervisor", "total", "ok", "nok", "fora", "taxa"]].copy()
        df_s.columns = ["Supervisor", "Total", "OK", "NOK", "Fora", "Taxa (%)"]
        df_s = df_s.sort_values("Taxa (%)")
        _titulo_aba(ws4, f"Resultado por Supervisor — {u['data_ref']}", len(df_s.columns))
        _write_header(ws4, df_s.columns.tolist(), 2)
        for ri, (_, row) in enumerate(df_s.iterrows(), 3):
            alt = ri % 2 == 0
            for ci, val in enumerate(row, 1):
                try:
                    if pd.isnull(val): val = None
                except: pass
                c = ws4.cell(ri, ci, val)
                c.font = Font(name="Calibri", size=10); c.alignment = _CTR; c.border = _BRD
                if alt: c.fill = PatternFill("solid", fgColor=C_ALT)
                if ci == 6: c.number_format = "0.0%"
                elif ci in (2, 3, 4, 5): c.number_format = "#,##0"
        _auto_width(ws4); ws4.freeze_panes = "A3"

    return StreamingResponse(
        _to_stream(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Triagem_{u['data_ref']}.xlsx"},
    )
